"""
Vista del docente - Plataforma Educativa IA
Versión Mejorada: Validaciones robustas + UI mejorada + Manejo de errores
CORRECCIONES:
- Esquema DB: Agregadas columnas 'level' y 'updated_at' para evitar crashes.
- Foro: Escapado de HTML para evitar inyección de código visual.
"""

import streamlit as st
import pandas as pd
import json
import base64
import time
import html  # Importante para corregir el error visual del foro
from datetime import datetime, date, timedelta
import traceback
from utils_notifications import notification_manager
import re

# Importación segura con fallback para evitar crasheos
try:
    from utils_ai import ai_generate_exam_from_text, display_pdf, extract_text_from_pdf, ai_manager
    AI_AVAILABLE = True
except ImportError as e:
    st.warning(f"⚠️ Algunas funciones de IA no disponibles: {e}")
    AI_AVAILABLE = False
    
    # Funciones de respaldo
    def ai_generate_exam_from_text(model, context_text, num_questions, num_options):
        return [{"question": "Función IA no disponible", "options": ["Opción 1", "Opción 2"], "correct_index": 0, "points": 5}]
    
    def extract_text_from_pdf(file_bytes):
        return "Error: Librería PDF no disponible. Instala pypdf con: pip install pypdf"
    
    def display_pdf(file_bytes):
        st.error("Visualizador PDF no disponible")

def init_teacher_state():
    """Inicializa estado del docente"""
    defaults = {
        'active_task_id': None,
        'active_task_title': "",
        'active_exam_review_id': None,
        'exam_draft': [],
        'current_module_id': None,
        'show_ai_exam_generator': False,
        'exam_generation_text': "",
        'exam_generation_file': None
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def check_teacher_schema(conn):
    """Verifica y actualiza esquema necesario para el docente"""
    try:
        c = conn.cursor()
        
        # Tabla Módulos
        c.execute("""CREATE TABLE IF NOT EXISTS modules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            order_index INTEGER DEFAULT 0,
            start_date DATE,
            end_date DATE,
            is_published INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE
        )""")
        
        # Tabla de foro mejorada
        c.execute("""CREATE TABLE IF NOT EXISTS forum_posts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id INTEGER NOT NULL,
            user_id TEXT NOT NULL,
            message TEXT NOT NULL,
            is_question INTEGER DEFAULT 0,
            is_resolved INTEGER DEFAULT 0,
            parent_id INTEGER,
            likes INTEGER DEFAULT 0,
            date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(username)
        )""")
        
        # Actualizaciones de columnas (safely)
        updates = [
            "ALTER TABLE course_materials ADD COLUMN module_id INTEGER",
            "ALTER TABLE tasks ADD COLUMN module_id INTEGER",
            "ALTER TABLE submissions ADD COLUMN teacher_feedback TEXT",
            "ALTER TABLE submissions ADD COLUMN file_name TEXT",
            "ALTER TABLE exam_questions ADD COLUMN question_type TEXT DEFAULT 'multiple_choice'",
            "ALTER TABLE exam_attempts ADD COLUMN details_json TEXT",
            "ALTER TABLE tasks ADD COLUMN allow_late_submissions INTEGER DEFAULT 1",
            "ALTER TABLE tasks ADD COLUMN max_attempts INTEGER DEFAULT 1",
            "ALTER TABLE exams ADD COLUMN is_published INTEGER DEFAULT 0",
            "ALTER TABLE exams ADD COLUMN passing_score INTEGER DEFAULT 60",
            # CORRECCIONES CRÍTICAS AQUÍ:
            "ALTER TABLE courses ADD COLUMN level TEXT DEFAULT 'Básico'",
            "ALTER TABLE courses ADD COLUMN updated_at TIMESTAMP"
        ]
        
        for sql in updates:
            try:
                c.execute(sql)
            except Exception as e:
                pass  # Ignorar si la columna ya existe
        
        conn.commit()
        return True
    except Exception as e:
        st.error(f"❌ Error en esquema de base de datos: {e}")
        return False

def validate_date_range(start_date, end_date):
    """Valida que las fechas sean válidas"""
    if not start_date or not end_date:
        return True, ""
    
    if end_date < start_date:
        return False, "La fecha de fin no puede ser anterior a la de inicio"
    
    return True, ""

def create_course_card(course, conn):
    """Crea una tarjeta visual para un curso"""
    desc = course.get('description', 'Sin descripción')
    desc = re.sub(r'<[^>]+>', '', desc).strip()
    if desc and len(desc) > 120:
        desc = desc[:120] + "..."
    
    if course.get('cover_image'):
        img_src = f"data:image/png;base64,{base64.b64encode(course['cover_image']).decode()}"
    else:
        img_src = "https://images.unsplash.com/photo-1501504905252-473c47e087f8?w=400&h=200&fit=crop"
    
    students_count = conn.execute(
        "SELECT COUNT(*) FROM enrollments WHERE course_id = ?", 
        (course['id'],)
    ).fetchone()[0]
    
    return f"""
    <div style="
        border: 1px solid #444; 
        border-radius: 15px; 
        overflow: hidden; 
        margin-bottom: 15px; 
        background-color: #1e1e1e; 
        height: 320px; 
        display: flex; 
        flex-direction: column;
        transition: transform 0.3s ease;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    ">
        <div style="height: 140px; overflow: hidden; background: #000;">
            <img src="{img_src}" style="width: 100%; height: 100%; object-fit: cover;">
        </div>
        <div style="padding: 15px; flex-grow: 1; display: flex; flex-direction: column;">
            <div style="color: #58a6ff; font-weight: bold; font-size: 0.8rem; margin-bottom: 5px;">
                {course.get('code', 'SIN-COD')}
            </div>
            <h3 style="margin: 0 0 8px 0; font-size: 1.1rem; color: white; line-height: 1.3;">
                {course.get('name', 'Curso sin nombre')}
            </h3>
            <p style="font-size: 0.85rem; color: #aaa; line-height: 1.4; flex-grow: 1; margin-bottom: 10px;">
                {desc}
            </p>
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <span style="font-size: 0.8rem; color: #888;">
                    👥 {students_count} alumnos
                </span>
                <span style="font-size: 0.8rem; color: #888; background: #2a2a2a; padding: 2px 8px; border-radius: 10px;">
                    {course.get('status', 'active')}
                </span>
            </div>
        </div>
    </div>
    """

def render_teacher_dashboard(conn, u):
    """Renderiza el dashboard del docente"""
    
    # Header con foto de perfil y saludo
    col1, col2 = st.columns([1, 8])
    
    with col1:
        # Mostrar avatar del docente
        if u.get('avatar'):
            avatar_b64 = base64.b64encode(u['avatar']).decode()
            st.markdown(f"""
            <div style="display: flex; justify-content: center; margin-top: 10px;">
                <img src="data:image/png;base64,{avatar_b64}" 
                     style="width: 80px; height: 80px; border-radius: 50%; 
                            border: 3px solid #58a6ff; object-fit: cover;">
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div style="display: flex; justify-content: center; margin-top: 10px;">
                <img src="https://cdn-icons-png.flaticon.com/512/847/847969.png" 
                     style="width: 80px; height: 80px; border-radius: 50%; 
                            border: 3px solid #58a6ff; object-fit: cover;">
            </div>
            """, unsafe_allow_html=True)
    
    with col2:
        st.title(f"👋 Panel Docente: {u.get('full_name', 'Docente')}")
        st.caption(f"👨‍🏫 Profesor • Última actividad: {datetime.now().strftime('%d/%m/%Y')}")
    
    st.markdown("---")
    
    st.subheader("📚 Mis Cursos")
    
    # Filtros
    col_filter1, col_filter2, col_filter3 = st.columns(3)
    filter_status = col_filter1.selectbox(
        "Filtrar por estado", 
        ["Todos", "Activos", "Borradores", "Archivados"],
        key="filter_status"
    )
    
    # Obtener cursos
    query = "SELECT * FROM courses WHERE teacher_id = ?"
    params = [u['username']]
    
    if filter_status != "Todos":
        status_map = {
            "Activos": "active",
            "Borradores": "draft", 
            "Archivados": "archived"
        }
        query += " AND status = ?"
        params.append(status_map[filter_status])
    
    query += " ORDER BY created_at DESC"
    
    courses_rows = conn.execute(query, tuple(params)).fetchall()
    courses = [dict(c) for c in courses_rows]
    
    if not courses:
        st.info("📭 No tienes cursos asignados. El administrador debe asignarte un curso.")
        return
    
    # Mostrar estadísticas
    total_courses = len(courses)
    active_courses = len([c for c in courses if c.get('status') == 'active'])
    draft_courses = len([c for c in courses if c.get('status') == 'draft'])
    
    col_stats1, col_stats2, col_stats3 = st.columns(3)
    col_stats1.metric("📚 Total Cursos", total_courses)
    col_stats2.metric("🟢 Activos", active_courses)
    col_stats3.metric("📝 Borradores", draft_courses)
    
    st.markdown("---")
    
    # Mostrar cursos en grid
    cols = st.columns(3)
    for i, course in enumerate(courses):
        with cols[i % 3]:
            st.markdown(create_course_card(course, conn), unsafe_allow_html=True)
            
            col_btn1, col_btn2 = st.columns(2)
            
            if col_btn1.button(
                "📝 Gestionar", 
                key=f"manage_{course['id']}", 
                width='stretch',
                help=f"Abrir {course['name']}"
            ):
                st.session_state.active_course = dict(course)
                st.session_state.view_mode = 'course'
                st.rerun()
            
            if col_btn2.button(
                "📊 Estadísticas", 
                key=f"stats_{course['id']}", 
                width='stretch',
                type="secondary"
            ):
                st.info(f"Estadísticas para {course['name']}")

def render_teacher_course_view(conn, model, u):
    """Renderiza la vista dentro de un curso"""
    c = st.session_state.active_course
    
    # Validación de seguridad
    if c is None or c.get('teacher_id') != u['username']:
        st.warning("⚠️ No tienes acceso a este curso. Volviendo al panel...")
        st.session_state.view_mode = 'dashboard'
        st.rerun()
        return
    
    # Barra superior con navegación
    col_back, col_title, col_actions = st.columns([1, 3, 2])
    
    with col_back:
        if st.button("← Volver a Cursos", type="tertiary", width='stretch'):
            st.session_state.view_mode = 'dashboard'
            st.rerun()
    
    with col_title:
        status_badge = {
            'active': '🟢 Activo',
            'draft': '📝 Borrador',
            'archived': '🗄️ Archivado'
        }.get(c.get('status', 'active'), '❓ Desconocido')
        
        st.title(f"📚 {c.get('name', 'Curso sin nombre')}")
        st.caption(f"🔖 {c.get('code', 'SIN-COD')} | 👨‍🏫 {u.get('full_name', 'Docente')} | {status_badge}")
    
    with col_actions:
        if st.button("🔄 Actualizar", key="refresh_course", type="secondary", width='stretch'):
            st.rerun()
    
    # Portada del curso
    if c.get('cover_image'):
        b64_img = base64.b64encode(c['cover_image']).decode()
        st.markdown(f"""
        <div style="
            width: 100%; 
            height: 200px; 
            overflow: hidden; 
            border-radius: 12px; 
            margin: 10px 0 20px 0;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        ">
            <img src="data:image/png;base64,{b64_img}" 
                 style="width: 100%; height: 100%; object-fit: cover;">
        </div>
        """, unsafe_allow_html=True)
    
    # Descripción del curso
    if c.get('description'):
        with st.expander("📖 Descripción del curso", expanded=False):
            st.markdown(c['description'])
    
    st.markdown("---")
    
    # ----------------------------------------------------------------------
    # PANELES DE GESTIÓN
    # ----------------------------------------------------------------------
    
    # 1. REVISOR DE EXÁMENES
    if st.session_state.get('active_exam_review_id'):
        render_exam_review_panel(conn, c, model)
        return  # No mostrar pestañas cuando se está revisando un examen
    
    # 2. GESTOR DE TAREAS
    if st.session_state.get('active_task_id'):
        render_task_manager_panel(conn, c)
        return  # No mostrar pestañas cuando se está gestionando una tarea
    
    # ======================================================================
    # PESTAÑAS PRINCIPALES DEL CURSO
    # ======================================================================
    tabs = st.tabs([
        "📚 Módulos", 
        "📝 Tareas", 
        "✅ Exámenes", 
        "💬 Foro", 
        "👥 Alumnos", 
        "📊 Calificaciones",
        "⚙️ Configuración"
    ])

    # --- TAB 1: MÓDULOS ---
    with tabs[0]:
        render_modules_tab(conn, model, c, u)
    
    # --- TAB 2: TAREAS ---
    with tabs[1]:
        render_tasks_tab(conn, c)
    
    # --- TAB 3: EXÁMENES ---
    with tabs[2]:
        render_exams_tab(conn, c)
    
    # --- TAB 4: FORO ---
    with tabs[3]:
        render_forum_tab(conn, c, u)
    
    # --- TAB 5: ALUMNOS ---
    with tabs[4]:
        render_students_tab(conn, c)
    
    # --- TAB 6: CALIFICACIONES ---
    with tabs[5]:
        render_grades_tab(conn, c)
    
    # --- TAB 7: CONFIGURACIÓN ---
    with tabs[6]:
        render_settings_tab(conn, c)

def render_exam_review_panel(conn, course, model):
    """Panel para revisar exámenes"""
    st.markdown("---")
    st.info(f"🔍 **Revisando Examen:** {st.session_state.get('active_task_title', 'Sin título')}")
    
    if st.button("❌ Cerrar Revisión", key="close_exam_rev"):
        st.session_state.active_exam_review_id = None
        st.rerun()
    
    exam_id = st.session_state.active_exam_review_id
    
    attempts_rows = conn.execute("""
        SELECT a.*, u.full_name FROM exam_attempts a 
        JOIN users u ON a.student_id = u.username 
        WHERE a.exam_id = ? ORDER BY a.end_time DESC
    """, (exam_id,)).fetchall()
    attempts = [dict(a) for a in attempts_rows]
    
    if not attempts:
        st.warning("📭 Aún no hay intentos de examen.")
    else:
        for attempt in attempts:
            # Verificar si hay preguntas pendientes
            has_pending = False
            if attempt.get('details_json'):
                try:
                    details = json.loads(attempt['details_json'])
                    has_pending = any(not d.get('graded', True) for d in details)
                except:
                    pass
            
            pending_badge = " 🔴 PENDIENTE" if has_pending else ""
            
            with st.expander(f"👤 {attempt.get('full_name', 'Estudiante')} - Nota: {attempt.get('score', 0)}{pending_badge}", expanded=has_pending):
                if attempt.get('details_json'):
                    try:
                        details = json.loads(attempt['details_json'])
                        
                        # Contar preguntas pendientes y por tipo
                        pending_questions = [d for d in details if not d.get('graded', True)]
                        open_questions = [d for d in details if d.get('type') == 'open_text']

                        if st.button("✏️ Editar Manualmente", key=f"manual_grade_{attempt['id']}",
                                     use_container_width=True):
                            st.info("👇 Ajusta las puntuaciones abajo y guarda los cambios")

                        # Mostrar estadísticas
                        st.markdown("---")
                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        total_points = sum(d.get('max_points', 0) for d in details)
                        current_score = sum(d.get('score', 0) for d in details)
                        col_stat1.metric("Nota Actual", f"{current_score}/{total_points}")
                        col_stat2.metric("Pendientes", len(pending_questions))
                        col_stat3.metric("Texto Abierto", len(open_questions))
                        
                        st.markdown("---")
                        
                        with st.form(f"review_attempt_{attempt['id']}"):
                            st.markdown("#### 📋 Detalle de Respuestas")
                            total_score = 0
                            
                            for idx, detail in enumerate(details):
                                is_pending = not detail.get('graded', True)
                                qtype = detail.get('type', 'multiple_choice')
                                status_icon = "🔴" if is_pending else "✅"
                                type_icon = "📝" if qtype == 'open_text' else "☑️"
                                
                                with st.container(border=True):
                                    st.markdown(f"{status_icon} {type_icon} **Pregunta {idx+1}** ({detail.get('max_points', 0)} puntos)")
                                    st.markdown(f"**{detail.get('question', 'Pregunta sin texto')}**")
                                    st.markdown(f"**Respuesta del estudiante:** {detail.get('answer', 'Sin responder')}")
                                    
                                    # Permitir editar TODAS las preguntas
                                    current_score = float(detail.get('score', 0))
                                    max_points = float(detail.get('max_points', 10))
                                    
                                    col_score, col_feedback = st.columns([1, 2])
                                    
                                    new_score = col_score.number_input(
                                        f"Puntuación (0-{max_points})",
                                        min_value=0.0,
                                        max_value=max_points,
                                        value=current_score,
                                        step=0.5,
                                        key=f"adj_{attempt['id']}_{detail.get('q_id', idx)}"
                                    )
                                    detail['score'] = new_score
                                    detail['graded'] = True
                                    
                                    # Mostrar feedback si existe
                                    if detail.get('needs_manual_review'):
                                        col_feedback.warning("⚠️ Requiere revisión manual — ajusta la puntuación")
                                    elif detail.get('ai_feedback'):
                                        col_feedback.info(f"💬 **Feedback IA:** {detail['ai_feedback']}")
                                    
                                    # Para opción múltiple, mostrar si era correcta originalmente
                                    if qtype == 'multiple_choice':
                                        original_correct = current_score > 0
                                        st.caption(f"✓ Respuesta {'correcta' if original_correct else 'incorrecta'} según clave de respuestas")
                                    
                                    total_score += float(detail.get('score', 0))
                            
                            st.markdown("---")
                            st.markdown(f"### 📊 Puntuación Total: **{total_score}** / {total_points}")
                            
                            col_save, col_cancel = st.columns([1, 1])
                            
                            if col_save.form_submit_button("💾 Guardar Corrección", type="primary", use_container_width=True):
                                updated_json = json.dumps(details)
                                conn.execute(
                                    "UPDATE exam_attempts SET score = ?, details_json = ? WHERE id = ?", 
                                    (total_score, updated_json, attempt['id'])
                                )
                                conn.commit()
                                st.success(f"✅ Nota actualizada: {total_score}")
                                time.sleep(1)
                                st.rerun()
                            
                            if col_cancel.form_submit_button("❌ Cancelar", type="secondary", use_container_width=True):
                                st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al cargar detalles: {str(e)}")
                else:
                    st.warning("📄 Este intento no tiene detalles guardados.")

def render_task_manager_panel(conn, course):
    """Panel para gestionar tareas"""
    st.markdown("---")
    
    # Botón para cerrar el gestor
    col_title, col_close = st.columns([4, 1])
    col_title.info(f"📝 **Gestionando Tarea:** {st.session_state.get('active_task_title', 'Sin título')}")
    
    if col_close.button("❌ Cerrar", key="close_task_man", type="secondary", use_container_width=True):
        if 'active_task_id' in st.session_state:
            del st.session_state.active_task_id
        if 'active_task_title' in st.session_state:
            del st.session_state.active_task_title
        st.rerun()
    
    task_id = st.session_state.get('active_task_id')
    
    if not task_id:
        st.error("⚠️ No se encontró la tarea. Volviendo...")
        time.sleep(1)
        st.rerun()
        return
    
    st.markdown("---")
    
    submissions_rows = conn.execute("""
        SELECT s.*, u.full_name FROM submissions s 
        JOIN users u ON s.student_id = u.username 
        WHERE s.task_id = ? ORDER BY s.submission_date DESC
    """, (task_id,)).fetchall()
    submissions = [dict(s) for s in submissions_rows]
    
    if not submissions:
        st.info("📭 Aún no hay entregas para esta tarea.")
    else:
        pending = [s for s in submissions if s.get('status') != 'graded']
        graded = [s for s in submissions if s.get('status') == 'graded']
        
        tab_pending, tab_graded = st.tabs([
            f"⏳ Pendientes ({len(pending)})", 
            f"✅ Calificadas ({len(graded)})"
        ])
        
        with tab_pending:
            if not pending:
                st.success("🎉 ¡Todas las tareas están calificadas!")
            else:
                for sub in pending:
                    with st.container(border=True):
                        st.markdown(f"### 👤 {sub.get('full_name', 'Estudiante')}")
                        if sub.get('submission_date'):
                            st.caption(f"📅 Entregado: {sub['submission_date']}")
                        
                        if sub.get('file_blob'):
                            file_name = sub.get('file_name', 'entrega.bin')
                            st.download_button(
                                f"📥 Descargar {file_name}",
                                sub['file_blob'],
                                file_name
                            )
                        
                        if sub.get('code'):
                            with st.expander("📝 Ver código"):
                                st.code(sub.get('code', ''))
                        
                        with st.form(f"grade_{sub['id']}"):
                            col_grade, col_max = st.columns(2)
                            grade = col_grade.number_input(
                                "Calificación", 
                                0.0, 100.0, 
                                value=float(sub.get('ai_grade', 0)) if sub.get('ai_grade') else 0.0, 
                                step=0.5,
                                key=f"grade_val_{sub['id']}"
                            )
                            
                            task_points = conn.execute(
                                "SELECT points FROM tasks WHERE id = ?", 
                                (task_id,)
                            ).fetchone()
                            max_points = task_points[0] if task_points else 100
                            col_max.metric("Puntos máx.", max_points)
                            
                            feedback = st.text_area(
                                "Feedback", 
                                value=sub.get('ai_feedback', '') if sub.get('ai_feedback') else '', 
                                placeholder="Escribe feedback personalizado para el estudiante...",
                                height=100,
                                key=f"feedback_{sub['id']}"
                            )
                            
                            if st.form_submit_button("📊 Calificar", type="primary"):
                                conn.execute(
                                    """
                                    UPDATE submissions 
                                    SET final_grade = ?, teacher_feedback = ?, status = 'graded', graded_date = ? 
                                    WHERE id = ?
                                    """, 
                                    (grade, feedback, datetime.now(), sub['id'])
                                )
                                conn.commit()
                                st.success("✅ Calificación guardada")
                                time.sleep(1)
                                st.rerun()
        
        with tab_graded:
            if not graded:
                st.info("ℹ️ No hay tareas calificadas aún.")
            else:
                for sub in graded:
                    st.markdown(f"""
                    <div style="
                        background: #1e1e1e; 
                        padding: 12px; 
                        border-radius: 8px; 
                        margin-bottom: 8px;
                        border-left: 4px solid #2ea043;
                    ">
                        <div style="font-weight: bold; font-size: 1rem;">
                            👤 {sub.get('full_name', 'Estudiante')}
                        </div>
                        <div style="display: flex; justify-content: space-between; margin-top: 5px;">
                            <span>📊 Nota: <strong>{sub.get('final_grade', 0)}</strong></span>
                            <span style="font-size: 0.8rem; color: #888;">
                                📅 {sub.get('graded_date', 'No especificado')}
                            </span>
                        </div>
                        <div style="margin-top: 5px; font-size: 0.9rem;">
                            💬 Feedback: {sub.get('teacher_feedback', 'Sin feedback')}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

def render_modules_tab(conn, model, course, user):
    """Renderiza la pestaña de módulos"""
    st.subheader("📁 Gestión de Módulos")
    
    # Crear nuevo módulo
    with st.form("create_module_form"):
        col_title, col_btn = st.columns([4, 1])
        new_mod_title = col_title.text_input(
            "Nuevo Módulo/Tema*", 
            placeholder="Ej: Semana 1 - Introducción a Python",
            help="Título descriptivo del módulo"
        )
        
        if col_btn.form_submit_button("➕ Crear", type="primary", width='stretch'):
            if new_mod_title.strip():
                try:
                    conn.execute(
                        "INSERT INTO modules (course_id, title) VALUES (?, ?)", 
                        (course['id'], new_mod_title.strip())
                    )
                    conn.commit()
                    st.success("✅ Módulo creado")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
            else:
                st.warning("⚠️ El título es obligatorio")
    
    st.divider()
    
    # Listar módulos existentes
    modules_rows = conn.execute(
        "SELECT * FROM modules WHERE course_id = ? ORDER BY order_index, id", 
        (course['id'],)
    ).fetchall()
    modules = [dict(m) for m in modules_rows]
    
    if not modules:
        st.info("📭 Crea tu primer módulo arriba para organizar el contenido.")
        return
    
    for module in modules:
        with st.expander(f"📌 {module.get('title', 'Módulo sin título')}", expanded=False):
            # Mostrar contenidos del módulo
            materials_rows = conn.execute(
                "SELECT * FROM course_materials WHERE module_id = ? ORDER BY order_index", 
                (module['id'],)
            ).fetchall()
            materials = [dict(m) for m in materials_rows]
            
            tasks_rows = conn.execute(
                "SELECT * FROM tasks WHERE module_id = ? ORDER BY due_date", 
                (module['id'],)
            ).fetchall()
            tasks = [dict(t) for t in tasks_rows]
            
            exams_rows = conn.execute(
                "SELECT * FROM exams WHERE module_id = ? ORDER BY created_at", 
                (module['id'],)
            ).fetchall()
            exams = [dict(e) for e in exams_rows]
            
            # Mostrar estadísticas rápidas
            col_mat, col_tasks, col_exams = st.columns(3)
            col_mat.metric("📄 Materiales", len(materials))
            col_tasks.metric("📝 Tareas", len(tasks))
            col_exams.metric("✅ Exámenes", len(exams))
            
            st.markdown("---")
            
            # Botones para Chat IA
            module_id = module['id']
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("🤖 Configurar Chat IA", key=f"config_chat_{module_id}", use_container_width=True):
                    st.session_state[f'show_chat_config_{module_id}'] = True
                    st.session_state[f'show_group_chat_teacher_{module_id}'] = False
                    st.rerun()
            with col_btn2:
                if st.button("👥 Ver Chat Grupal", key=f"view_group_chat_{module_id}", use_container_width=True, type="secondary"):
                    st.session_state[f'show_group_chat_teacher_{module_id}'] = True
                    st.session_state[f'show_chat_config_{module_id}'] = False
                    st.rerun()

            # Mostrar configuración de chat si está activada
            if st.session_state.get(f'show_chat_config_{module_id}'):
                render_module_chat_config(conn, module_id, course['id'], model)
                
                if st.button("❌ Cerrar configuración", key=f"close_chat_config_{module_id}"):
                    st.session_state[f'show_chat_config_{module_id}'] = False
                    st.rerun()
                
                st.markdown("---")

            # Mostrar chat grupal del módulo (vista del profesor)
            if st.session_state.get(f'show_group_chat_teacher_{module_id}'):
                from views_student import render_group_chat_interface
                render_group_chat_interface(conn, module_id, user['username'], 'teacher', module['title'], model)

                if st.button("❌ Cerrar chat grupal", key=f"close_group_chat_teacher_{module_id}"):
                    st.session_state[f'show_group_chat_teacher_{module_id}'] = False
                    st.rerun()

                st.markdown("---")
            
            # Contenidos existentes
            if materials or tasks or exams:
                st.markdown("#### 📚 Contenidos existentes")
                
                for material in materials:
                    col_icon, col_info, col_del = st.columns([0.5, 4, 0.5])
                    col_icon.markdown("📄")
                    col_info.write(f"**{material.get('title', 'Sin título')}**")
                    if col_del.button("🗑️", key=f"del_mat_{material['id']}", help="Eliminar material"):
                        conn.execute("DELETE FROM course_materials WHERE id = ?", (material['id'],))
                        conn.commit()
                        st.rerun()
                
                for task in tasks:
                    col_icon, col_info, col_rev, col_del = st.columns([0.5, 3, 1, 0.5])
                    col_icon.markdown("📝")
                    col_info.write(f"**{task.get('title', 'Sin título')}**")
                    if col_rev.button("📋 Revisar", key=f"rev_task_m_{task['id']}"):
                        st.session_state.active_task_id = task['id']
                        st.session_state.active_task_title = task.get('title', 'Tarea')
                        st.rerun()
                    if col_del.button("🗑️", key=f"del_task_{task['id']}", help="Eliminar tarea"):
                        conn.execute("DELETE FROM tasks WHERE id = ?", (task['id'],))
                        conn.commit()
                        st.rerun()
                
                for exam in exams:
                    col_icon, col_info, col_rev, col_del = st.columns([0.5, 3, 1, 0.5])
                    col_icon.markdown("✅")
                    col_info.write(f"**{exam.get('title', 'Sin título')}**")
                    if col_rev.button("📊 Revisar", key=f"rev_exam_m_{exam['id']}"):
                        st.session_state.active_exam_review_id = exam['id']
                        st.session_state.active_task_title = exam.get('title', 'Examen')
                        st.rerun()
                    if col_del.button("🗑️", key=f"del_exam_{exam['id']}", help="Eliminar examen"):
                        conn.execute("DELETE FROM exams WHERE id = ?", (exam['id'],))
                        conn.commit()
                        st.rerun()
            
            st.markdown("---")
            st.markdown(f"#### ➕ Agregar contenido a: {module.get('title', 'Módulo')}")
            
            sub_tabs = st.tabs(["📄 Material", "📝 Tarea", "✅ Examen"])
            
            with sub_tabs[0]:  # Material
                with st.form(f"add_material_{module['id']}"):
                    mat_title = st.text_input("Título del Material*")
                    mat_desc = st.text_area("Descripción")
                    mat_file = st.file_uploader("Archivo (PDF, imágenes, documentos)", type=['pdf', 'png', 'jpg', 'jpeg', 'docx', 'txt'])
                    
                    if st.form_submit_button("📤 Subir Material"):
                        if mat_title:
                            file_blob = mat_file.getvalue() if mat_file else None
                            # Determinar tipo correcto según el constraint de la BD
                            if mat_file:
                                if mat_file.name.endswith('.pdf'):
                                    file_type = 'pdf'
                                else:
                                    file_type = 'text'  # Cambiado de 'document' a 'text'
                            else:
                                file_type = 'text'
                            
                            conn.execute(
                                """
                                INSERT INTO course_materials 
                                (course_id, module_id, title, content_text, content_blob, type, date) 
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                                """, 
                                (course['id'], module['id'], mat_title, mat_desc, file_blob, file_type, datetime.now())
                            )
                            conn.commit()
                            st.success("✅ Material agregado")
                            st.rerun()
                        else:
                            st.warning("⚠️ El título es obligatorio")
            
            with sub_tabs[1]:  # Tarea
                with st.form(f"add_task_{module['id']}"):
                    task_title = st.text_input("Título de la Tarea*")
                    task_points = st.number_input("Nota máxima (1-10)*", 1, 10, 10)
                    task_due = st.date_input("Fecha de Entrega*", value=date.today() + timedelta(days=7))
                    task_type = st.selectbox("Tipo de Entrega*", ["file", "code"], format_func=lambda x: "📎 Archivo" if x == "file" else "💻 Código")
                    task_desc = st.text_area("Instrucciones*", placeholder="Describe detalladamente la tarea...", height=120)
                    
                    col_opt1, col_opt2 = st.columns(2)
                    allow_late = col_opt1.checkbox("¿Permitir entregas tardías?", value=True)
                    max_attempts = col_opt2.number_input("Intentos máx.", 1, 5, 1)
                    
                    if st.form_submit_button("📝 Crear Tarea", type="primary"):
                        if task_title and task_desc:
                            try:
                                conn.execute("""
                                    INSERT INTO tasks 
                                    (course_id, module_id, title, description, due_date, points, submission_type, created_by, allow_late_submissions, max_attempts) 
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    course['id'], module['id'], task_title, task_desc, 
                                    task_due, task_points, task_type, user['username'], 
                                    1 if allow_late else 0, max_attempts
                                ))
                                conn.commit()
                                st.success("✅ Tarea creada exitosamente")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error al guardar tarea: {str(e)}")
                        else:
                            st.warning("⚠️ El título y las instrucciones son obligatorios.")
            
            with sub_tabs[2]:  # Examen
                render_exam_creation_section(conn, model, course, module)
            
            # Botón para eliminar módulo
            st.markdown("---")
            if st.button("🗑️ Eliminar Módulo", key=f"del_mod_{module['id']}", type="secondary"):
                if st.checkbox("⚠️ Confirmar eliminación del módulo y todo su contenido"):
                    conn.execute("DELETE FROM modules WHERE id = ?", (module['id'],))
                    conn.commit()
                    st.success("✅ Módulo eliminado")
                    time.sleep(1)
                    st.rerun()

def render_exam_creation_section(conn, model, course, module):
    """Renderiza la sección MEJORADA para crear exámenes con IA y configuración flexible"""
    
    st.markdown("### 📝 Crear Nuevo Examen")
    
    # Configuración básica del examen
    with st.container(border=True):
        st.markdown("#### ⚙️ Configuración General")
        
        col1, col2 = st.columns(2)
        exam_title = col1.text_input("Título del Examen*", key=f"exam_title_{module['id']}")
        exam_duration = col2.number_input("Duración (minutos)*", 10, 300, 60, key=f"exam_dur_{module['id']}")
        
        col3, col4 = st.columns(2)
        passing_score = col3.slider("Nota mínima de aprobación", 1, 10, 6, key=f"pass_score_{module['id']}")
        total_exam_points = 10  # Escala fija del 1 al 10
        col4.info("📊 Nota máxima: **10** (escala del 1 al 10)")
    
    # Configuración de preguntas
    with st.container(border=True):
        st.markdown("#### 📊 Configuración de Preguntas")
        
        col_q1, col_q2 = st.columns(2)
        num_multiple_choice = col_q1.number_input(
            "Preguntas de opción múltiple", 
            0, 100, 5, 
            key=f"num_mc_{module['id']}",
            help="Sin límite de preguntas"
        )
        num_open_text = col_q2.number_input(
            "Preguntas de desarrollo/texto", 
            0, 100, 2, 
            key=f"num_open_{module['id']}",
            help="Sin límite de preguntas"
        )
        
        total_questions = num_multiple_choice + num_open_text
        st.info(f"📝 Total de preguntas: **{total_questions}**")
        
        if total_questions == 0:
            st.warning("⚠️ Debes tener al menos 1 pregunta")
    
    # Método de creación
    st.markdown("---")
    exam_mode = st.radio(
        "Método de creación:", 
        ["🤖 Generar con IA (Recomendado)", "✍️ Manual"], 
        key=f"exam_mode_{module['id']}", 
        horizontal=True
    )
    
    if exam_mode == "🤖 Generar con IA (Recomendado)":
        st.markdown("#### 🤖 Generación Automática con IA")
        st.info("📚 La IA generará preguntas basadas en el contenido que proporciones")
        
        # Opciones de entrada
        input_method = st.radio(
            "Fuente del contenido:",
            ["📝 Texto", "📄 PDF nuevo", "📂 PDF del módulo", "📚 Ambos (texto + PDF)"],
            horizontal=True,
            key=f"input_method_{module['id']}"
        )
        
        text_input = ""
        pdf_text = ""
        
        if input_method in ["📝 Texto", "📚 Ambos (texto + PDF)"]:
            text_input = st.text_area(
                "Pega el contenido aquí", 
                key=f"text_{module['id']}", 
                height=200,
                placeholder="Pega el contenido del tema, apuntes, o material de estudio..."
            )
        
        if input_method == "📂 PDF del módulo":
            # Obtener PDFs subidos en la semana en este módulo o curso
            week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
            course_id_row = conn.execute("SELECT course_id FROM modules WHERE id = ?", (module['id'],)).fetchone()
            course_id_val = course_id_row['course_id'] if course_id_row else None
            
            if course_id_val:
                recent_pdfs = conn.execute("""
                    SELECT id, title, date, type FROM course_materials
                    WHERE (module_id = ? OR course_id = ?)
                    AND (type = 'pdf' OR type LIKE '%pdf%' OR title LIKE '%.pdf')
                    AND date >= ?
                    ORDER BY date DESC
                """, (module['id'], course_id_val, week_ago)).fetchall()
                recent_pdfs = [dict(r) for r in recent_pdfs]
            else:
                recent_pdfs = []
            
            if recent_pdfs:
                st.success(f"📂 {len(recent_pdfs)} PDF(s) encontrados en los últimos 7 días")
                pdf_selected = st.selectbox(
                    "Selecciona un PDF:",
                    [f"📄 {p['title']} ({p['date'][:10] if p['date'] else 'sin fecha'})" for p in recent_pdfs],
                    key=f"pdf_sel_{module['id']}"
                )
                # Obtener el índice seleccionado
                pdf_idx = [f"📄 {p['title']} ({p['date'][:10] if p['date'] else 'sin fecha'})" for p in recent_pdfs].index(pdf_selected)
                selected_pdf_row = recent_pdfs[pdf_idx]
                
                # Leer el blob del PDF
                pdf_blob_row = conn.execute(
                    "SELECT content_blob FROM course_materials WHERE id = ?",
                    (selected_pdf_row['id'],)
                ).fetchone()
                
                if pdf_blob_row and pdf_blob_row['content_blob']:
                    with st.spinner("📖 Extrayendo texto del PDF..."):
                        pdf_text = extract_text_from_pdf(pdf_blob_row['content_blob'])
                        if pdf_text and "Error" not in pdf_text:
                            st.success(f"✅ PDF cargado: {len(pdf_text)} caracteres extraídos")
                        else:
                            st.error("⚠️ No se pudo extraer texto de este PDF")
                            pdf_text = ""
                else:
                    st.warning("⚠️ Este material no tiene contenido PDF adjunto")
            else:
                st.info("📭 No hay PDFs subidos en los últimos 7 días en este módulo. Sube uno nuevo.")
                # Fallback: permitir subir uno nuevo
                pdf_file = st.file_uploader(
                    "📄 Sube un documento PDF", 
                    type="pdf", 
                    key=f"pdf_fallback_{module['id']}"
                )
                if pdf_file:
                    with st.spinner("📖 Extrayendo texto del PDF..."):
                        pdf_text = extract_text_from_pdf(pdf_file.getvalue())
                        if "Error" not in pdf_text:
                            st.success(f"✅ PDF procesado: {len(pdf_text)} caracteres")
        
        if input_method in ["📄 PDF nuevo", "📚 Ambos (texto + PDF)"]:
            pdf_file = st.file_uploader(
                "📄 Sube un documento PDF", 
                type="pdf", 
                key=f"pdf_{module['id']}"
            )
            if pdf_file:
                with st.spinner("📖 Extrayendo texto del PDF..."):
                    pdf_text = extract_text_from_pdf(pdf_file.getvalue())
                    if "Error" not in pdf_text and "PyPDF" not in pdf_text:
                        st.success(f"✅ PDF procesado: {len(pdf_text)} caracteres extraídos")
                    else:
                        st.error(f"⚠️ {pdf_text}")
                        pdf_text = ""
        
        # Configuración adicional para IA
        with st.expander("⚙️ Configuración Avanzada de IA"):
            col_ai1, col_ai2 = st.columns(2)
            difficulty_level = col_ai1.selectbox(
                "Nivel de dificultad",
                ["Fácil", "Medio", "Difícil"],
                index=1,
                key=f"difficulty_{module['id']}"
            )
            num_options_mc = col_ai2.slider(
                "Opciones por pregunta (múltiple)",
                2, 6, 4,
                key=f"num_opts_{module['id']}"
            )
        
        # Botón para generar
        if st.button("✨ Generar Examen con IA", key=f"gen_ai_{module['id']}", type="primary", use_container_width=True):
            if not AI_AVAILABLE:
                st.error("⚠️ Funciones de IA no disponibles. Verifica la configuración.")
                return
            
            combined_text = (text_input + "\n\n" + pdf_text).strip()
            
            if len(combined_text) < 100:
                st.error("⚠️ El contenido es muy corto. Proporciona al menos 100 caracteres.")
                return
            
            with st.spinner("🤖 Generando preguntas con IA... Esto puede tomar un momento."):
                try:
                    generated_questions = []
                    
                    # Generar preguntas de opción múltiple
                    if num_multiple_choice > 0:
                        with st.status(f"Generando {num_multiple_choice} preguntas de opción múltiple..."):
                            mc_questions = ai_generate_exam_from_text(
                                model, 
                                combined_text, 
                                num_multiple_choice, 
                                num_options_mc
                            )
                            for q in mc_questions:
                                q['type'] = 'multiple_choice'
                                q['points'] = round(total_exam_points / total_questions, 1)
                            generated_questions.extend(mc_questions)
                            st.write(f"✅ {len(mc_questions)} preguntas de opción múltiple generadas")
                    
                    # Generar preguntas de desarrollo
                    if num_open_text > 0:
                        with st.status(f"Generando {num_open_text} preguntas de desarrollo...") as status:
                            try:
                                prompt = f"""
Genera EXACTAMENTE {num_open_text} preguntas de desarrollo/texto abierto basadas en este contenido:

{combined_text[:8000]}

Nivel de dificultad: {difficulty_level}

Las preguntas deben:
- Requerir respuestas elaboradas (2-3 párrafos mínimo)
- Evaluar comprensión profunda del tema
- Ser específicas al contenido proporcionado
- Variar en enfoque (análisis, aplicación, síntesis, evaluación)

IMPORTANTE: Responde SOLO con un array JSON válido, sin texto adicional, sin markdown.

Formato EXACTO:
[
    {{
        "question": "Primera pregunta de desarrollo aquí",
        "points": 10,
        "difficulty": "medium"
    }},
    {{
        "question": "Segunda pregunta de desarrollo aquí",
        "points": 10,
        "difficulty": "medium"
    }}
]

Genera EXACTAMENTE {num_open_text} preguntas en formato JSON.
"""
                                
                                # Usar la función mejorada de IA con reintentos
                                from utils_ai import AIManager
                                ai_manager = AIManager(st.secrets.get("GEMINI_API_KEY", ""))
                                
                                response_text = ai_manager.call_with_retry(
                                    prompt, 
                                    max_retries=3, 
                                    max_output_tokens=3000, 
                                    temperature=0.7
                                )
                                
                                if not response_text or "Error" in response_text:
                                    st.warning(f"⚠️ No se pudieron generar preguntas de desarrollo: {response_text}")
                                    status.update(label="⚠️ Error al generar preguntas de desarrollo", state="error")
                                else:
                                    # Limpiar respuesta
                                    cleaned = response_text.replace("```json", "").replace("```", "").strip()
                                    
                                    # Extraer JSON con múltiples estrategias
                                    import re
                                    qs_open = None
                                    
                                    # Estrategia 1: Buscar array JSON
                                    json_match = re.search(r'\[.*\]', cleaned, re.DOTALL)
                                    if json_match:
                                        try:
                                            qs_open = json.loads(json_match.group(0))
                                        except json.JSONDecodeError:
                                            # Estrategia 2: Limpiar y reintentar
                                            json_str = json_match.group(0)
                                            json_str = re.sub(r',\s*}', '}', json_str)
                                            json_str = re.sub(r',\s*]', ']', json_str)
                                            try:
                                                qs_open = json.loads(json_str)
                                            except:
                                                pass
                                    
                                    if qs_open and isinstance(qs_open, list) and len(qs_open) > 0:
                                        # Procesar preguntas
                                        for q in qs_open:
                                            q['type'] = 'open_text'
                                            q['options'] = []
                                            q['correct_index'] = -1
                                            if 'points' not in q:
                                                q['points'] = round(total_exam_points / total_questions, 1)
                                            else:
                                                q['points'] = round(total_exam_points / total_questions, 1)
                                        
                                        generated_questions.extend(qs_open)
                                        st.write(f"✅ {len(qs_open)} preguntas de desarrollo generadas")
                                        status.update(label=f"✅ {len(qs_open)} preguntas de desarrollo generadas", state="complete")
                                    else:
                                        st.warning("⚠️ No se pudieron parsear las preguntas de desarrollo")
                                        st.code(cleaned[:500], language="text")
                                        status.update(label="⚠️ Error al parsear preguntas", state="error")
                                        
                            except Exception as e:
                                st.error(f"❌ Error generando preguntas de desarrollo: {str(e)}")
                                status.update(label=f"❌ Error: {str(e)[:50]}", state="error")
                    
                    # Agregar al borrador
                    if generated_questions:
                        # Forzar exactamente el número de preguntas solicitado
                        expected_total = num_multiple_choice + num_open_text
                        if len(generated_questions) < expected_total:
                            # Faltan preguntas — generar las faltantes
                            missing = expected_total - len(generated_questions)
                            mc_missing = max(0, num_multiple_choice - len([q for q in generated_questions if q.get('type') == 'multiple_choice']))
                            open_missing = max(0, num_open_text - len([q for q in generated_questions if q.get('type') == 'open_text']))
                            
                            extra_prompt = f"""
Genera EXACTAMENTE {missing} preguntas adicionales sobre este contenido:
{combined_text[:4000]}

IMPORTANTE: Responde SOLO con JSON array válido.
{f'Genera {mc_missing} preguntas de opción múltiple con {num_options_mc} opciones.' if mc_missing > 0 else ''}
{f'Genera {open_missing} preguntas de desarrollo.' if open_missing > 0 else ''}

Formato para opción múltiple:
[{{"question":"...", "options":["A","B","C","D"], "correct_index":0, "type":"multiple_choice"}}]
Formato para desarrollo:
[{{"question":"...", "options":[], "correct_index":-1, "type":"open_text"}}]
"""
                            try:
                                extra_response = ai_manager.call_with_retry(extra_prompt, max_retries=2, max_output_tokens=2000, temperature=0.8)
                                if extra_response:
                                    import re
                                    json_match = re.search(r'\[.*\]', extra_response, re.DOTALL)
                                    if json_match:
                                        extra_qs = json.loads(json_match.group(0))
                                        for q in extra_qs:
                                            if 'type' not in q:
                                                q['type'] = 'multiple_choice'
                                            if 'options' not in q:
                                                q['options'] = []
                                            if 'correct_index' not in q:
                                                q['correct_index'] = -1
                                            q['points'] = round(total_exam_points / expected_total, 1)
                                        generated_questions.extend(extra_qs)
                            except Exception:
                                pass
                        
                        # Recortar si hay demasiadas
                        mc_qs = [q for q in generated_questions if q.get('type') == 'multiple_choice'][:num_multiple_choice]
                        open_qs = [q for q in generated_questions if q.get('type') == 'open_text'][:num_open_text]
                        final_questions = mc_qs + open_qs
                        
                        # Recalcular puntos para que sumen exactamente 10
                        n = len(final_questions)
                        if n > 0:
                            pts_each = round(10.0 / n, 2)
                            for q in final_questions:
                                q['points'] = pts_each
                        
                        st.session_state.exam_draft = final_questions  # Reemplazar, no agregar
                        st.success(f"🎉 {len(final_questions)} preguntas generadas ({num_multiple_choice} opción múltiple + {num_open_text} desarrollo)")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("❌ No se generaron preguntas. Intenta de nuevo.")
                        
                except Exception as e:
                    st.error(f"❌ Error al generar preguntas: {str(e)}")
    
    else:  # Modo Manual
        st.markdown("#### ✍️ Creación Manual de Preguntas")
        
        with st.form(f"manual_exam_{module['id']}"):
            q_type = st.selectbox("Tipo de Pregunta", ["Opción Múltiple", "Texto Abierto/Desarrollo"])
            q_text = st.text_area("Pregunta*", height=100)
            q_points = st.number_input("Puntos*", 1, 50, 5)
            
            options = []
            correct_idx = -1
            
            if q_type == "Opción Múltiple":
                num_opts = st.slider("Número de opciones", 2, 6, 4)
                st.markdown("**Opciones:**")
                
                for i in range(num_opts):
                    letter = chr(65 + i)  # A, B, C, ...
                    opt = st.text_input(f"Opción {letter}*", key=f"opt_{module['id']}_{i}")
                    options.append(opt)
                
                correct_letter = st.selectbox("Respuesta correcta", [chr(65 + i) for i in range(num_opts)])
                correct_idx = ord(correct_letter) - 65
            else:
                st.info("💡 Las preguntas de texto abierto serán calificadas manualmente o con ayuda de IA")
            
            if st.form_submit_button("➕ Agregar Pregunta al Borrador", type="primary", use_container_width=True):
                if q_text and (q_type != "Opción Múltiple" or all(options)):
                    q_data = {
                        "question": q_text, 
                        "options": options, 
                        "correct_index": correct_idx, 
                        "points": q_points, 
                        "type": "open_text" if q_type == "Texto Abierto/Desarrollo" else "multiple_choice"
                    }
                    st.session_state.exam_draft.append(q_data)
                    st.success("✅ Pregunta agregada al borrador")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.warning("⚠️ Completa todos los campos obligatorios")
    
    # Mostrar borrador actual con opciones de edición
    if st.session_state.exam_draft:
        st.markdown("---")
        st.markdown(f"### 📝 Borrador de Examen ({len(st.session_state.exam_draft)} preguntas)")
        
        total_points = sum(q.get('points', 0) for q in st.session_state.exam_draft)
        mc_count = len([q for q in st.session_state.exam_draft if q.get('type') == 'multiple_choice'])
        open_count = len([q for q in st.session_state.exam_draft if q.get('type') == 'open_text'])
        
        col_stat1, col_stat2, col_stat3 = st.columns(3)
        col_stat1.metric("📊 Puntos totales", f"{total_points:.1f}")
        col_stat2.metric("☑️ Opción múltiple", mc_count)
        col_stat3.metric("✍️ Desarrollo", open_count)
        
        # Opción para cambiar puntos de todas las preguntas
        with st.expander("⚙️ Ajustar Puntuación"):
            st.markdown("**Cambiar puntos de todas las preguntas:**")
            col_pts1, col_pts2 = st.columns([3, 1])
            new_points_all = col_pts1.number_input(
                "Nuevos puntos para TODAS las preguntas",
                1, 50, 5,
                key=f"new_pts_all_{module['id']}"
            )
            if col_pts2.button("Aplicar a todas", key=f"apply_all_{module['id']}"):
                for q in st.session_state.exam_draft:
                    q['points'] = new_points_all
                st.success(f"✅ Todas las preguntas ahora valen {new_points_all} puntos")
                st.rerun()
        
        # Mostrar cada pregunta con opciones de edición
        for idx, question in enumerate(st.session_state.exam_draft):
            with st.expander(
                f"**{idx+1}.** {question.get('question', 'Pregunta sin texto')[:80]}... "
                f"({question.get('points', 0)} pts)",
                expanded=False
            ):
                st.markdown(f"**Pregunta completa:**")
                st.write(question.get('question', 'Sin texto'))
                
                q_type_label = "✍️ Desarrollo" if question.get('type') == 'open_text' else "☑️ Opción Múltiple"
                st.caption(f"Tipo: {q_type_label}")
                
                if question.get('type') == 'multiple_choice':
                    st.markdown("**Opciones:**")
                    for i, option in enumerate(question.get('options', [])):
                        prefix = "✅" if i == question.get('correct_index') else "⚪"
                        st.write(f"{prefix} {chr(65+i)}. {option}")
                
                # --- EDICIÓN COMPLETA ---
                edit_key = f"editing_{module['id']}_{idx}"
                if st.button("✏️ Editar esta pregunta", key=f"edit_btn_{module['id']}_{idx}"):
                    st.session_state[edit_key] = True
                
                if st.session_state.get(edit_key):
                    st.markdown("**✏️ Editando pregunta:**")
                    
                    new_q_text = st.text_area(
                        "Texto de la pregunta:",
                        value=question.get('question', ''),
                        key=f"edit_q_text_{module['id']}_{idx}",
                        height=100
                    )
                    
                    new_pts = st.number_input(
                        "Puntos:",
                        min_value=0.1, max_value=10.0,
                        value=float(question.get('points', 1.0)),
                        step=0.1,
                        key=f"edit_pts2_{module['id']}_{idx}"
                    )
                    
                    new_options = question.get('options', []).copy()
                    new_correct = question.get('correct_index', 0)
                    
                    if question.get('type') == 'multiple_choice':
                        st.markdown("**Opciones:**")
                        for i in range(len(new_options)):
                            new_options[i] = st.text_input(
                                f"Opción {chr(65+i)}:",
                                value=new_options[i],
                                key=f"edit_opt_{module['id']}_{idx}_{i}"
                            )
                        correct_letters = [chr(65+i) for i in range(len(new_options))]
                        selected_letter = st.selectbox(
                            "Respuesta correcta:",
                            correct_letters,
                            index=max(0, question.get('correct_index', 0)),
                            key=f"edit_correct_{module['id']}_{idx}"
                        )
                        new_correct = ord(selected_letter) - 65
                    
                    col_save_edit, col_cancel_edit = st.columns(2)
                    if col_save_edit.button("💾 Guardar cambios", key=f"save_edit_{module['id']}_{idx}", type="primary"):
                        st.session_state.exam_draft[idx]['question'] = new_q_text
                        st.session_state.exam_draft[idx]['points'] = new_pts
                        if question.get('type') == 'multiple_choice':
                            st.session_state.exam_draft[idx]['options'] = new_options
                            st.session_state.exam_draft[idx]['correct_index'] = new_correct
                        st.session_state.pop(edit_key, None)
                        st.success("✅ Pregunta actualizada")
                        st.rerun()
                    if col_cancel_edit.button("❌ Cancelar", key=f"cancel_edit_{module['id']}_{idx}"):
                        st.session_state.pop(edit_key, None)
                        st.rerun()
                
                # Eliminar
                if st.button("🗑️ Eliminar pregunta", key=f"del_q_{module['id']}_{idx}", type="secondary"):
                    st.session_state.exam_draft.pop(idx)
                    st.session_state.pop(edit_key, None)
                    st.rerun()
        
        # Botones de acción
        st.markdown("---")
        col_save, col_clear = st.columns([3, 1])
        
        if col_clear.button("🗑️ Limpiar Borrador", key=f"clear_{module['id']}", type="secondary", use_container_width=True):
            st.session_state.exam_draft = []
            st.rerun()
        
        if col_save.button("💾 Guardar Examen Completo", key=f"save_exam_{module['id']}", type="primary", use_container_width=True):
            if not exam_title:
                st.error("⚠️ El título del examen es obligatorio")
                return
            
            if len(st.session_state.exam_draft) == 0:
                st.error("⚠️ Debes tener al menos 1 pregunta")
                return
            
            try:
                # Guardar examen
                conn.execute(
                    """
                    INSERT INTO exams 
                    (course_id, module_id, title, duration_minutes, passing_score) 
                    VALUES (?, ?, ?, ?, ?)
                    """, 
                    (course['id'], module['id'], exam_title, exam_duration, passing_score)
                )
                exam_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                
                # Guardar preguntas
                for q in st.session_state.exam_draft:
                    conn.execute(
                        """
                        INSERT INTO exam_questions 
                        (exam_id, question, options_json, correct_index, points, question_type) 
                        VALUES (?, ?, ?, ?, ?, ?)
                        """, 
                        (
                            exam_id, 
                            q['question'], 
                            json.dumps(q.get('options', [])), 
                            q.get('correct_index', -1), 
                            q.get('points', 5), 
                            q.get('type', 'multiple_choice')
                        )
                    )
                
                conn.commit()
                st.session_state.exam_draft = []
                st.success("🎉 Examen guardado exitosamente")
                
                # Notificar a estudiantes
                try:
                    from utils_notifications import NotificationManager
                    notif_manager = NotificationManager(conn)
                    notif_manager.create_exam_notification(course['id'], exam_title)
                except:
                    pass
                
                time.sleep(1.5)
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error al guardar examen: {str(e)}")
    else:
        st.info("📝 El borrador está vacío. Agrega preguntas para comenzar.")

def render_module_chat_config(conn, module_id, course_id, model):
    """
    Renderiza la configuración de chat IA para un módulo
    Se integra en la pestaña de Módulos
    """
    from utils_chat_ai import ModuleChatManager
    from utils_ai import extract_text_from_pdf
    
    ai_manager = model  # El model ya es una instancia de AIManager
    chat_manager = ModuleChatManager(conn, ai_manager)
    
    st.markdown("### 🤖 Configuración de Chat IA")
    
    # Verificar si ya existe configuración
    existing_content = chat_manager.get_chat_content(module_id)
    
    if existing_content:
        st.success("✅ Chat IA configurado para este módulo")
        
        with st.expander("📄 Ver contenido actual", expanded=False):
            preview_text = existing_content['content_text'][:500]
            if len(existing_content['content_text']) > 500:
                preview_text += "..."
            
            st.text_area(
                "Contenido de contexto:",
                value=preview_text,
                height=150,
                disabled=True,
                key=f"preview_{module_id}"
            )
            st.caption(f"Tipo: {existing_content['content_type']} | Creado: {existing_content['created_at']}")
        
        # Mostrar preguntas sugeridas
        questions = chat_manager.get_suggested_questions(module_id)
        if questions:
            st.markdown("**💡 Preguntas sugeridas actuales:**")
            for i, q in enumerate(questions, 1):
                st.markdown(f"{i}. {q['question_text']}")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Actualizar contenido", key=f"update_chat_{module_id}", use_container_width=True):
                st.session_state[f'editing_chat_{module_id}'] = True
                st.rerun()
        
        with col2:
            if st.button("🗑️ Eliminar chat IA", key=f"delete_chat_{module_id}", type="secondary", use_container_width=True):
                if chat_manager.delete_module_chat(module_id):
                    st.success("Chat IA eliminado")
                    time.sleep(1)
                    st.rerun()
    
    # Formulario de configuración
    if not existing_content or st.session_state.get(f'editing_chat_{module_id}'):
        st.markdown("#### Configurar contenido del chat")
        
        # Radio button FUERA del formulario para que funcione correctamente
        content_type = st.radio(
            "Tipo de contenido:",
            ["Texto", "PDF", "📂 PDF del módulo"],
            horizontal=True,
            key=f"content_type_{module_id}"
        )
        
        # Variables para almacenar el contenido extraído
        extracted_text = None
        extracted_file_name = None
        
        # File uploader FUERA del formulario (solo para PDF nuevo)
        if content_type == "PDF":
            uploaded_file = st.file_uploader(
                "Subir archivo PDF:",
                type=['pdf'],
                help="El texto del PDF será extraído automáticamente",
                key=f"pdf_upload_{module_id}"
            )
            
            if uploaded_file:
                extracted_file_name = uploaded_file.name
                pdf_bytes = uploaded_file.getvalue()
                
                with st.spinner("Extrayendo texto del PDF..."):
                    try:
                        extracted_text = extract_text_from_pdf(pdf_bytes)
                        
                        if extracted_text:
                            st.session_state[f'extracted_text_{module_id}'] = extracted_text
                            st.session_state[f'extracted_file_name_{module_id}'] = extracted_file_name
                            st.success(f"✅ Texto extraído: {len(extracted_text)} caracteres")
                            with st.expander("Vista previa del texto extraído"):
                                preview = extracted_text[:500]
                                if len(extracted_text) > 500:
                                    preview += "..."
                                st.text_area("Vista previa", value=preview, height=150, disabled=True, key=f"preview_extracted_{module_id}", label_visibility="collapsed")
                        else:
                            st.error("❌ No se pudo extraer texto del PDF")
                    except Exception as e:
                        st.error(f"❌ Error al extraer texto: {str(e)}")
        
        # Opción de PDF del módulo (últimos 7 días)
        elif content_type == "📂 PDF del módulo":
            from datetime import timedelta
            week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
            
            # Obtener course_id del módulo
            course_id_row = conn.execute("SELECT course_id FROM modules WHERE id = ?", (module_id,)).fetchone()
            course_id_val = course_id_row['course_id'] if course_id_row else None
            
            if course_id_val:
                recent_pdfs = conn.execute("""
                    SELECT id, title, date FROM course_materials
                    WHERE (module_id = ? OR course_id = ?)
                    AND (type = 'pdf' OR type LIKE '%pdf%' OR title LIKE '%.pdf')
                    AND date >= ?
                    ORDER BY date DESC
                """, (module_id, course_id_val, week_ago)).fetchall()
                recent_pdfs = [dict(r) for r in recent_pdfs]
            else:
                recent_pdfs = []
            
            if recent_pdfs:
                st.success(f"📂 {len(recent_pdfs)} PDF(s) disponibles de los últimos 7 días")
                pdf_options = [f"📄 {p['title']} ({(p['date'] or '')[:10]})" for p in recent_pdfs]
                selected_label = st.selectbox("Selecciona un PDF:", pdf_options, key=f"pdf_module_sel_{module_id}")
                pdf_idx = pdf_options.index(selected_label)
                selected_pdf = recent_pdfs[pdf_idx]
                
                if st.button("📖 Cargar PDF seleccionado", key=f"load_pdf_{module_id}", type="secondary"):
                    blob_row = conn.execute(
                        "SELECT content_blob FROM course_materials WHERE id = ?", (selected_pdf['id'],)
                    ).fetchone()
                    if blob_row and blob_row['content_blob']:
                        with st.spinner("Extrayendo texto..."):
                            extracted_text = extract_text_from_pdf(blob_row['content_blob'])
                            if extracted_text and "Error" not in extracted_text:
                                st.session_state[f'extracted_text_{module_id}'] = extracted_text
                                st.session_state[f'extracted_file_name_{module_id}'] = selected_pdf['title']
                                st.success(f"✅ PDF cargado: {len(extracted_text)} caracteres")
                                st.rerun()
                            else:
                                st.error("⚠️ No se pudo extraer texto de este PDF")
                    else:
                        st.warning("⚠️ Este material no tiene contenido PDF adjunto")
                
                # Mostrar si ya está cargado
                if f'extracted_text_{module_id}' in st.session_state:
                    fname = st.session_state.get(f'extracted_file_name_{module_id}', 'PDF')
                    txt = st.session_state[f'extracted_text_{module_id}']
                    st.info(f"📄 PDF listo: {fname} ({len(txt)} caracteres)")
                    extracted_text = txt
                    extracted_file_name = fname
            else:
                st.info("📭 No hay PDFs subidos en los últimos 7 días. Sube uno nuevo con la opción 'PDF'.")
        
        # Ahora el formulario
        with st.form(f"chat_config_{module_id}"):
            content_text = None
            file_name = None
            
            if content_type == "Texto":
                content_text = st.text_area(
                    "Contenido educativo:",
                    height=200,
                    placeholder="Escribe o pega el contenido del módulo aquí...",
                    help="Este texto será el contexto para las respuestas de la IA",
                    key=f"content_text_{module_id}"
                )
            else:
                # Usar el texto extraído del PDF guardado en session_state (aplica para PDF nuevo y PDF del módulo)
                if f'extracted_text_{module_id}' in st.session_state:
                    content_text = st.session_state[f'extracted_text_{module_id}']
                    file_name = st.session_state.get(f'extracted_file_name_{module_id}')
                    st.info(f"📄 PDF listo para guardar: {file_name} ({len(content_text)} caracteres)")
                else:
                    if content_type == "📂 PDF del módulo":
                        st.warning("👆 Selecciona y carga un PDF del módulo arriba")
                    else:
                        st.warning("👆 Primero sube un archivo PDF arriba")
            
            col_submit, col_cancel = st.columns(2)
            
            with col_submit:
                submitted = st.form_submit_button("💾 Guardar configuración", type="primary", use_container_width=True)
            
            with col_cancel:
                cancelled = st.form_submit_button("❌ Cancelar", use_container_width=True)
            
            if submitted:
                if not content_text or len(content_text) < 50:
                    st.error("⚠️ El contenido es demasiado corto (mínimo 50 caracteres)")
                else:
                    with st.spinner("Configurando chat IA..."):
                        result = chat_manager.configure_module_chat(
                            module_id=module_id,
                            content_type='pdf' if content_type == "PDF" else 'text',
                            content=content_text,
                            file_name=file_name
                        )
                    
                    if result['success']:
                        st.success("✅ Chat IA configurado exitosamente")
                        num_questions = len(result.get('questions', []))
                        st.info(f"💡 Se generaron {num_questions} preguntas sugeridas")
                        
                        if 'warning' in result:
                            st.warning(result['warning'])
                        
                        # Limpiar session_state
                        if f'editing_chat_{module_id}' in st.session_state:
                            del st.session_state[f'editing_chat_{module_id}']
                        if f'extracted_text_{module_id}' in st.session_state:
                            del st.session_state[f'extracted_text_{module_id}']
                        if f'extracted_file_name_{module_id}' in st.session_state:
                            del st.session_state[f'extracted_file_name_{module_id}']
                        
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error(f"❌ Error: {result.get('error', 'Error desconocido')}")
            
            if cancelled:
                # Limpiar session_state
                if f'editing_chat_{module_id}' in st.session_state:
                    del st.session_state[f'editing_chat_{module_id}']
                if f'extracted_text_{module_id}' in st.session_state:
                    del st.session_state[f'extracted_text_{module_id}']
                if f'extracted_file_name_{module_id}' in st.session_state:
                    del st.session_state[f'extracted_file_name_{module_id}']
                st.rerun()

def render_tasks_tab(conn, course):
    """Renderiza la pestaña de tareas"""
    st.subheader("📋 Todas las Tareas")
    
    all_tasks_rows = conn.execute("""
        SELECT t.*, m.title as module_title FROM tasks t 
        LEFT JOIN modules m ON t.module_id = m.id 
        WHERE t.course_id = ? ORDER BY t.due_date DESC
    """, (course['id'],)).fetchall()
    all_tasks = [dict(t) for t in all_tasks_rows]
    
    if not all_tasks:
        st.info("📭 No hay tareas creadas.")
        return
    
    # Estadísticas
    total_tasks = len(all_tasks)
    graded_tasks = len([t for t in all_tasks if conn.execute(
        "SELECT COUNT(*) FROM submissions WHERE task_id = ? AND status = 'graded'", 
        (t['id'],)
    ).fetchone()[0] > 0])
    
    col_stats1, col_stats2 = st.columns(2)
    col_stats1.metric("📝 Total Tareas", total_tasks)
    col_stats2.metric("✅ Tareas calificadas", graded_tasks)
    
    st.markdown("---")
    
    # Listar tareas
    for task in all_tasks:
        module_tag = f"[{task.get('module_title', 'General')}]"
        
        with st.expander(f"{module_tag} {task.get('title', 'Sin título')}", expanded=False):
            col_info1, col_info2 = st.columns(2)
            col_info1.write(f"**📅 Vence:** {task.get('due_date', 'No especificada')}")
            col_info2.write(f"**💯 Puntos:** {task.get('points', 0)}")
            
            if task.get('allow_late_submissions', 1):
                col_info1.success("🕒 Entregas tardías permitidas")
            else:
                col_info1.error("⛔ Entregas tardías NO permitidas")
            
            col_info2.write(f"**🔄 Intentos máx:** {task.get('max_attempts', 1)}")
            
            st.markdown("**📝 Descripción:**")
            st.markdown(task.get('description', 'Sin descripción'))
            
            # Estadísticas de entregas
            submissions_count = conn.execute(
                "SELECT COUNT(*) FROM submissions WHERE task_id = ?", 
                (task['id'],)
            ).fetchone()[0]
            
            graded_count = conn.execute(
                "SELECT COUNT(*) FROM submissions WHERE task_id = ? AND status = 'graded'", 
                (task['id'],)
            ).fetchone()[0]
            
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            col_stats1.metric("📤 Entregas", submissions_count)
            col_stats2.metric("✅ Calificadas", graded_count)
            col_stats3.metric("⏳ Pendientes", submissions_count - graded_count)
            
            # Botones: gestionar y eliminar
            col_btn1, col_btn2 = st.columns([2, 1])
            
            if col_btn1.button("📋 Gestionar Entregas", key=f"manage_task_{task['id']}", type="primary", use_container_width=True):
                st.session_state.active_task_id = task['id']
                st.session_state.active_task_title = task.get('title', 'Tarea')
                st.rerun()
            
            if col_btn2.button("🗑️ Eliminar Tarea", key=f"task_tab_del_{task['id']}", type="secondary", use_container_width=True):
                st.session_state[f"confirm_del_task_{task['id']}"] = True
            
            if st.session_state.get(f"confirm_del_task_{task['id']}"):
                st.warning(f"¿Eliminar la tarea **{task.get('title')}** y todas sus entregas?")
                col_yes, col_no = st.columns(2)
                if col_yes.button("✅ Sí, eliminar", key=f"task_tab_yes_{task['id']}", type="primary"):
                    try:
                        conn.execute("DELETE FROM submissions WHERE task_id = ?", (task['id'],))
                        conn.execute("DELETE FROM tasks WHERE id = ?", (task['id'],))
                        conn.commit()
                        st.session_state.pop(f"confirm_del_task_{task['id']}", None)
                        st.success("✅ Tarea eliminada")
                        time.sleep(0.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
                if col_no.button("❌ Cancelar", key=f"task_tab_no_{task['id']}"):
                    st.session_state.pop(f"confirm_del_task_{task['id']}", None)
                    st.rerun()

def render_exams_tab(conn, course):
    """Renderiza la pestaña de exámenes"""
    st.subheader("✅ Todos los Exámenes")
    
    all_exams_rows = conn.execute("""
        SELECT e.*, m.title as module_title FROM exams e 
        LEFT JOIN modules m ON e.module_id = m.id 
        WHERE e.course_id = ? ORDER BY e.created_at DESC
    """, (course['id'],)).fetchall()
    all_exams = [dict(e) for e in all_exams_rows]
    
    if not all_exams:
        st.info("📭 No hay exámenes creados.")
        return
    
    # Estadísticas
    total_exams = len(all_exams)
    published_exams = len([e for e in all_exams if e.get('is_published', 0) == 1])
    
    col_stats1, col_stats2 = st.columns(2)
    col_stats1.metric("📝 Total Exámenes", total_exams)
    col_stats2.metric("📢 Publicados", published_exams)
    
    st.markdown("---")
    
    # Listar exámenes
    for exam in all_exams:
        module_tag = f"[{exam.get('module_title', 'General')}]"
        
        with st.container(border=True):
            col_info, col_stats, col_action = st.columns([3, 2, 1])
            
            col_info.markdown(f"**{module_tag} {exam.get('title', 'Sin título')}**")
            col_info.caption(f"⏱️ {exam.get('duration_minutes', 0)} min | 📊 Aprobación: {exam.get('passing_score', 60)}%")
            
            # Estadísticas del examen
            attempts_count = conn.execute(
                "SELECT COUNT(*) FROM exam_attempts WHERE exam_id = ?", 
                (exam['id'],)
            ).fetchone()[0]
            
            avg_score = conn.execute(
                "SELECT AVG(score) FROM exam_attempts WHERE exam_id = ?", 
                (exam['id'],)
            ).fetchone()[0]
            
            col_stats.metric("📊 Intentos", attempts_count)
            if avg_score:
                col_stats.metric("📈 Promedio", f"{avg_score:.1f}")
            
            # Acciones
            if col_action.button("📋 Revisar", key=f"review_exam_{exam['id']}"):
                st.session_state.active_exam_review_id = exam['id']
                st.session_state.active_task_title = exam.get('title', 'Examen')
                st.rerun()
            
            col_publish, col_delete = st.columns(2)
            
            publish_status = "Publicar" if not exam.get('is_published') else "Despublicar"
            if col_publish.button(publish_status, key=f"publish_{exam['id']}", type="secondary"):
                new_status = 0 if exam.get('is_published') else 1
                conn.execute(
                    "UPDATE exams SET is_published = ? WHERE id = ?", 
                    (new_status, exam['id'])
                )
                conn.commit()
                st.rerun()
            
            if col_delete.button("🗑️", key=f"delete_exam_{exam['id']}"):
                st.session_state[f"confirm_del_exam_{exam['id']}"] = True
            
            if st.session_state.get(f"confirm_del_exam_{exam['id']}"):
                st.warning(f"¿Eliminar examen '{exam.get('title')}'? Esta acción también eliminará todos los intentos asociados.")
                col_yes, col_no = st.columns(2)
                if col_yes.button("✅ Sí, eliminar", key=f"confirm_yes_exam_{exam['id']}", type="primary"):
                    try:
                        # Eliminar intentos primero (por si no hay CASCADE)
                        conn.execute("DELETE FROM exam_attempts WHERE exam_id = ?", (exam['id'],))
                        conn.execute("DELETE FROM exam_questions WHERE exam_id = ?", (exam['id'],))
                        conn.execute("DELETE FROM exams WHERE id = ?", (exam['id'],))
                        conn.commit()
                        st.session_state.pop(f"confirm_del_exam_{exam['id']}", None)
                        st.success("✅ Examen eliminado")
                        time.sleep(0.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al eliminar: {e}")
                if col_no.button("❌ Cancelar", key=f"cancel_del_exam_{exam['id']}"):
                    st.session_state.pop(f"confirm_del_exam_{exam['id']}", None)
                    st.rerun()

def render_forum_tab(conn, course, user):
    """Renderiza la pestaña del foro (CORREGIDA - escape de HTML)"""
    st.markdown("#### 💬 Foro del Curso")
    
    # Publicar mensaje
    with st.form("teacher_post_form", clear_on_submit=True):
        message = st.text_area("Mensaje:", height=100, placeholder="Comparte información, anuncios o responde preguntas...")
        is_question = st.checkbox("Marcar como pregunta", value=False)
        
        col_submit, col_clear = st.columns([3, 1])
        
        if col_submit.form_submit_button("📤 Publicar", type="primary", width='stretch'):
            if message.strip():
                conn.execute(
                    """
                    INSERT INTO forum_posts (course_id, user_id, message, is_question, date) 
                    VALUES (?, ?, ?, ?, ?)
                    """, 
                    (course['id'], user['username'], message.strip(), 1 if is_question else 0, datetime.now())
                )
                conn.commit()
                st.success("✅ Mensaje publicado")
                st.rerun()
            else:
                st.warning("⚠️ Escribe un mensaje antes de publicar.")
        
        if col_clear.form_submit_button("🗑️ Limpiar", type="secondary", width='stretch'):
            st.rerun()
    
    st.divider()
    
    # Mostrar mensajes
    posts_rows = conn.execute("""
        SELECT f.*, u.full_name, u.avatar, u.role 
        FROM forum_posts f 
        JOIN users u ON f.user_id = u.username 
        WHERE f.course_id = ? 
        ORDER BY f.date DESC 
        LIMIT 20
    """, (course['id'],)).fetchall()
    posts = [dict(p) for p in posts_rows]
    
    if not posts:
        st.info("💭 No hay mensajes en el foro. ¡Sé el primero en publicar!")
        return
    
    # Filtros
    col_filter1, col_filter2 = st.columns(2)
    show_questions = col_filter1.checkbox("Mostrar solo preguntas", value=False)
    show_resolved = col_filter2.checkbox("Mostrar resueltas", value=True)
    
    filtered_posts = posts
    if show_questions:
        filtered_posts = [p for p in filtered_posts if p.get('is_question')]
    if not show_resolved:
        filtered_posts = [p for p in filtered_posts if not p.get('is_resolved')]
    
    st.caption(f"📄 Mostrando {len(filtered_posts)} de {len(posts)} mensajes")
    
    for post in filtered_posts:
        # Determinar estilo
        is_teacher = post.get('role') == 'teacher'
        is_own = post['user_id'] == user['username']
        
        bg_color = "#2b2d42" if is_teacher else "#1e1e1e"
        border_color = "#ffcc00" if is_teacher else "transparent"
        
        # Avatar
        if post.get('avatar'):
            b64_av = base64.b64encode(post['avatar']).decode()
            img_html = f'<img src="data:image/png;base64,{b64_av}" style="width:40px;height:40px;border-radius:50%;object-fit:cover;">'
        else:
            img_html = '<img src="https://cdn-icons-png.flaticon.com/512/847/847969.png" style="width:40px;height:40px;border-radius:50%;">'
        
        # Iconos de estado
        status_icons = ""
        if post.get('is_question'):
            status_icons += "❓ "
        if post.get('is_resolved'):
            status_icons += "✅ "
        
        role_badge = " 👨‍🏫" if is_teacher else " 🎓"
        
        # Obtener y limpiar el mensaje
        import re
        import html as html_module
        message_content = post.get('message', '')
        
        # Eliminar etiquetas HTML (incluyendo multilinea y self-closing)
        clean_message = re.sub(r'<[^>]*>', '', message_content, flags=re.DOTALL)
        
        # Decodificar entidades HTML (&amp; &lt; etc.)
        clean_message = html_module.unescape(clean_message)
        
        # Limpiar espacios/líneas vacías excesivas
        clean_message = re.sub(r'\n{3,}', '\n\n', clean_message).strip()
        
        # Convertir saltos de línea a <br> para HTML
        safe_message = html_module.escape(clean_message).replace('\n', '<br>')
        
        # Si el mensaje está vacío, mostrar mensaje por defecto
        if not safe_message.strip():
            safe_message = '<em style="color: #888;">Mensaje sin contenido</em>'
        
        post_html = (
            f'<div style="display:flex;gap:10px;margin-top:10px;padding:15px;'
            f'background:{bg_color};border-radius:10px;border-left:4px solid {border_color};align-items:flex-start;">'
            f'{img_html}'
            f'<div style="flex-grow:1;">'
            f'<div style="font-weight:bold;font-size:0.9rem;">'
            f'{post.get("full_name","Usuario")}{role_badge} '
            f'<span style="color:#aaa;font-weight:normal;font-size:0.8rem;">• {post.get("date","")}</span>'
            f' {status_icons} {"(Tú)" if is_own else ""}'
            f'</div>'
            f'<div style="margin-top:8px;font-size:0.95rem;line-height:1.5;">{safe_message}</div>'
            f'</div></div>'
        )
        st.markdown(post_html, unsafe_allow_html=True)
        
        # Acciones para el profesor
        if post.get('is_question') and not post.get('is_resolved') and is_teacher:
            col_resolve, col_reply = st.columns([1, 2])
            if col_resolve.button("✅ Marcar como resuelta", key=f"resolve_{post['id']}"):
                conn.execute("UPDATE forum_posts SET is_resolved = 1 WHERE id = ?", (post['id'],))
                conn.commit()
                st.rerun()
            
            if col_reply.button("💬 Responder", key=f"reply_{post['id']}"):
                st.info("Funcionalidad de respuesta en desarrollo")

def render_students_tab(conn, course):
    """Renderiza la pestaña de alumnos"""
    st.subheader("👥 Alumnos Inscritos")
    
    # Estadísticas
    total_students = conn.execute(
        "SELECT COUNT(*) FROM enrollments WHERE course_id = ?", 
        (course['id'],)
    ).fetchone()[0]
    
    active_students = conn.execute("""
        SELECT COUNT(DISTINCT s.student_id) 
        FROM submissions s 
        JOIN enrollments e ON s.student_id = e.student_id 
        WHERE e.course_id = ? AND s.submission_date >= DATE('now', '-30 days')
    """, (course['id'],)).fetchone()[0]
    
    col_stats1, col_stats2 = st.columns(2)
    col_stats1.metric("👥 Total Inscritos", total_students)
    col_stats2.metric("📈 Activos (30d)", active_students)
    
    # Obtener alumnos
    students_rows = conn.execute("""
        SELECT u.*, e.enrollment_date 
        FROM users u 
        JOIN enrollments e ON u.username = e.student_id 
        WHERE e.course_id = ? AND u.role = 'student' 
        ORDER BY u.full_name
    """, (course['id'],)).fetchall()
    students = [dict(s) for s in students_rows]
    
    if not students:
        st.info("📭 No hay estudiantes inscritos en este curso.")
        return
    
    # Mostrar tabla de alumnos
    df_data = []
    for student in students:
        # Obtener estadísticas del estudiante
        submissions_count = conn.execute(
            "SELECT COUNT(*) FROM submissions WHERE student_id = ? AND task_id IN (SELECT id FROM tasks WHERE course_id = ?)", 
            (student['username'], course['id'])
        ).fetchone()[0]
        
        avg_grade = conn.execute("""
            SELECT AVG(final_grade) 
            FROM submissions 
            WHERE student_id = ? AND status = 'graded' 
            AND task_id IN (SELECT id FROM tasks WHERE course_id = ?)
        """, (student['username'], course['id'])).fetchone()[0]
        
        df_data.append({
            "Nombre": student.get('full_name', ''),
            "Usuario": student.get('username', ''),
            "Email": student.get('email', ''),
            "Inscrito": student.get('enrollment_date', ''),
            "Entregas": submissions_count,
            "Promedio": f"{avg_grade:.1f}" if avg_grade else "-"
        })
    
    df = pd.DataFrame(df_data)
    st.dataframe(
        df, 
        width='stretch', 
        hide_index=True,
        column_config={
            "Nombre": st.column_config.TextColumn(width="medium"),
            "Usuario": st.column_config.TextColumn(width="small"),
            "Email": st.column_config.TextColumn(width="large"),
            "Entregas": st.column_config.NumberColumn(width="small"),
            "Promedio": st.column_config.NumberColumn(width="small", format="%.1f")
        }
    )
    
    # Botones de acción
    col_export, col_invite = st.columns(2)
    
    if col_export.button("📥 Exportar Lista (CSV)"):
        csv = df.to_csv(index=False)
        st.download_button(
            label="⬇️ Descargar CSV",
            data=csv,
            file_name=f"alumnos_{course['code']}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
    
    if col_invite.button("📧 Enviar Invitaciones"):
        st.info("Funcionalidad de invitaciones en desarrollo")

def render_grades_tab(conn, course):
    """Renderiza la pestaña de calificaciones"""
    st.subheader("📊 Libro de Calificaciones")
    
    # Obtener estudiantes
    students_rows = conn.execute("""
        SELECT u.username, u.full_name 
        FROM users u 
        JOIN enrollments e ON u.username = e.student_id 
        WHERE e.course_id = ? AND u.role = 'student' 
        ORDER BY u.full_name
    """, (course['id'],)).fetchall()
    students = [dict(s) for s in students_rows]
    
    if not students:
        st.info("📭 No hay estudiantes inscritos.")
        return
    
    # Obtener tareas y exámenes
    tasks_list = [dict(t) for t in conn.execute(
        "SELECT id, title FROM tasks WHERE course_id = ?", 
        (course['id'],)
    ).fetchall()]
    
    exams_list = [dict(e) for e in conn.execute(
        "SELECT id, title FROM exams WHERE course_id = ?", 
        (course['id'],)
    ).fetchall()]
    
    # Construir tabla de calificaciones
    data = []
    for student in students:
        row = {"Estudiante": student['full_name']}
        
        # Calificaciones de tareas
        for task in tasks_list:
            grade_row = conn.execute(
                "SELECT final_grade FROM submissions WHERE task_id = ? AND student_id = ? AND status = 'graded'", 
                (task['id'], student['username'])
            ).fetchone()
            grade = dict(grade_row) if grade_row else {}
            grade_value = grade.get('final_grade')
            # Convert to string to avoid mixed types in DataFrame
            row[f"📝 {task['title']}"] = str(grade_value) if grade_value is not None else '-'
        
        # Calificaciones de exámenes
        for exam in exams_list:
            score_row = conn.execute(
                "SELECT score FROM exam_attempts WHERE exam_id = ? AND student_id = ?", 
                (exam['id'], student['username'])
            ).fetchone()
            score = dict(score_row) if score_row else {}
            score_value = score.get('score')
            # Convert to string to avoid mixed types in DataFrame
            row[f"✅ {exam['title']}"] = str(score_value) if score_value is not None else '-'
        
        data.append(row)
    
    if data:
        df = pd.DataFrame(data)
        
        # Calcular promedios
        if len(df.columns) > 1:
            numeric_cols = [col for col in df.columns if col != 'Estudiante']
            df['Promedio'] = df[numeric_cols].apply(
                lambda row: pd.to_numeric(row, errors='coerce').mean(), axis=1
            ).round(1)
        
        st.dataframe(
            df, 
            width='stretch',
            hide_index=True,
            column_config={
                "Estudiante": st.column_config.TextColumn(width="medium"),
                "Promedio": st.column_config.NumberColumn(width="small", format="%.1f")
            }
        )
        
        # Botones de exportación
        col_csv, col_excel = st.columns(2)
        
        if col_csv.button("📥 Exportar CSV"):
            csv = df.to_csv(index=False)
            st.download_button(
                label="⬇️ Descargar CSV",
                data=csv,
                file_name=f"calificaciones_{course['code']}_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
        
        if col_excel.button("📊 Exportar Excel"):
            try:
                import openpyxl
                from openpyxl.styles import (
                    Font, PatternFill, Alignment, Border, Side, GradientFill
                )
                from openpyxl.utils import get_column_letter
                import io

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Calificaciones"

                # ── Paleta de colores ──────────────────────────────────────
                COLOR_HEADER_BG   = "1F3864"   # Azul oscuro — encabezado columnas
                COLOR_HEADER_FONT = "FFFFFF"   # Blanco
                COLOR_TITLE_BG    = "2E75B6"   # Azul medio — fila título curso
                COLOR_SUBHDR_BG   = "D6E4F7"   # Azul muy claro — fila subtítulo info
                COLOR_ROW_ODD     = "F2F7FF"   # Filas impares
                COLOR_ROW_EVEN    = "FFFFFF"   # Filas pares
                COLOR_PROMEDIO_BG = "E2EFDA"   # Verde claro — columna promedio
                COLOR_ALTO        = "C6EFCE"   # Verde — nota alta (≥70)
                COLOR_MEDIO       = "FFEB9C"   # Amarillo — nota media (50–69)
                COLOR_BAJO        = "FFC7CE"   # Rojo claro — nota baja (<50)
                COLOR_BORDER      = "B8CCE4"   # Borde azul suave

                def border(style="thin"):
                    s = Side(style=style, color=COLOR_BORDER)
                    return Border(left=s, right=s, top=s, bottom=s)

                def fill(hex_color):
                    return PatternFill("solid", fgColor=hex_color)

                # ── Fila 1: Título del curso ──────────────────────────────
                total_cols = len(df.columns)
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=total_cols)
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = f"Libro de Calificaciones — {course.get('name', 'Curso')}"
                title_cell.font = Font(name="Calibri", bold=True, size=14,
                                       color=COLOR_HEADER_FONT)
                title_cell.fill = fill(COLOR_TITLE_BG)
                title_cell.alignment = Alignment(horizontal="center",
                                                 vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 28

                # ── Fila 2: Metadatos del curso ───────────────────────────
                ws.merge_cells(start_row=2, start_column=1,
                               end_row=2, end_column=total_cols)
                meta_cell = ws.cell(row=2, column=1)
                meta_cell.value = (
                    f"Código: {course.get('code', '-')}   |   "
                    f"Docente: {course.get('teacher_name') or course.get('teacher_id', '-')}   |   "
                    f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                )
                meta_cell.font = Font(name="Calibri", italic=True, size=10,
                                      color="444444")
                meta_cell.fill = fill(COLOR_SUBHDR_BG)
                meta_cell.alignment = Alignment(horizontal="center",
                                                vertical="center")
                ws.row_dimensions[2].height = 18

                # ── Fila 3: Encabezados de columnas ───────────────────────
                HEADER_ROW = 3
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=HEADER_ROW, column=col_idx)
                    # Limpiar emojis para Excel (algunos clientes los muestran mal)
                    clean_name = col_name
                    cell.value = clean_name
                    cell.font = Font(name="Calibri", bold=True, size=10,
                                     color=COLOR_HEADER_FONT)
                    cell.fill = fill(COLOR_HEADER_BG)
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=True)
                    cell.border = border()
                ws.row_dimensions[HEADER_ROW].height = 36

                # ── Filas de datos ────────────────────────────────────────
                for row_idx, (_, data_row) in enumerate(df.iterrows(), start=1):
                    excel_row = HEADER_ROW + row_idx
                    row_bg = COLOR_ROW_ODD if row_idx % 2 == 1 else COLOR_ROW_EVEN

                    for col_idx, col_name in enumerate(df.columns, start=1):
                        cell = ws.cell(row=excel_row, column=col_idx)
                        raw_val = data_row[col_name]

                        # Determinar valor y formato
                        if col_name == "Estudiante":
                            cell.value = str(raw_val)
                            cell.font = Font(name="Calibri", bold=True, size=10)
                            cell.fill = fill(row_bg)
                            cell.alignment = Alignment(horizontal="left",
                                                       vertical="center",
                                                       wrap_text=True)
                        elif col_name == "Promedio":
                            num = pd.to_numeric(raw_val, errors='coerce')
                            cell.value = float(num) if pd.notna(num) else None
                            cell.number_format = "0.0"
                            cell.font = Font(name="Calibri", bold=True, size=10)
                            # Color según rango
                            if pd.notna(num):
                                if num >= 70:
                                    cell.fill = fill(COLOR_ALTO)
                                elif num >= 50:
                                    cell.fill = fill(COLOR_MEDIO)
                                else:
                                    cell.fill = fill(COLOR_BAJO)
                            else:
                                cell.fill = fill(COLOR_PROMEDIO_BG)
                            cell.alignment = Alignment(horizontal="center",
                                                       vertical="center")
                        else:
                            # Notas de tareas / exámenes
                            if raw_val == '-' or raw_val is None:
                                cell.value = "-"
                                cell.fill = fill(row_bg)
                            else:
                                num = pd.to_numeric(raw_val, errors='coerce')
                                if pd.notna(num):
                                    cell.value = float(num)
                                    cell.number_format = "0.0"
                                    # Color semáforo
                                    if num >= 70:
                                        cell.fill = fill(COLOR_ALTO)
                                    elif num >= 50:
                                        cell.fill = fill(COLOR_MEDIO)
                                    else:
                                        cell.fill = fill(COLOR_BAJO)
                                else:
                                    cell.value = str(raw_val)
                                    cell.fill = fill(row_bg)
                            cell.alignment = Alignment(horizontal="center",
                                                       vertical="center")
                            cell.font = Font(name="Calibri", size=10)

                        cell.border = border()
                    ws.row_dimensions[excel_row].height = 18

                # ── Ajuste automático de anchos de columna ────────────────
                for col_idx, col_name in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    if col_name == "Estudiante":
                        ws.column_dimensions[col_letter].width = 28
                    elif col_name == "Promedio":
                        ws.column_dimensions[col_letter].width = 12
                    else:
                        # Ancho basado en el título (máx 22, mín 10)
                        ws.column_dimensions[col_letter].width = max(
                            10, min(22, len(col_name) * 0.85)
                        )

                # ── Fijar encabezados al hacer scroll ─────────────────────
                ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=2)

                # ── Hoja de leyenda ───────────────────────────────────────
                ws_legend = wb.create_sheet("Leyenda")
                legend_data = [
                    ("Color", "Significado", "Rango"),
                    ("Verde",  "Nota aprobada",  "≥ 70"),
                    ("Amarillo", "Nota en riesgo", "50 – 69"),
                    ("Rojo",   "Nota reprobada", "< 50"),
                    ("-",      "Sin calificación", "—"),
                ]
                legend_fills = [
                    fill(COLOR_HEADER_BG),
                    fill(COLOR_ALTO),
                    fill(COLOR_MEDIO),
                    fill(COLOR_BAJO),
                    fill(COLOR_ROW_ODD),
                ]
                legend_fonts = [
                    Font(bold=True, color=COLOR_HEADER_FONT),
                    Font(bold=False),
                    Font(bold=False),
                    Font(bold=False),
                    Font(bold=False),
                ]
                for r_idx, (row_data, r_fill, r_font) in enumerate(
                    zip(legend_data, legend_fills, legend_fonts), start=1
                ):
                    for c_idx, val in enumerate(row_data, start=1):
                        lc = ws_legend.cell(row=r_idx, column=c_idx, value=val)
                        lc.fill = r_fill
                        lc.font = r_font
                        lc.alignment = Alignment(horizontal="center")
                        lc.border = border()
                ws_legend.column_dimensions["A"].width = 14
                ws_legend.column_dimensions["B"].width = 22
                ws_legend.column_dimensions["C"].width = 14

                # ── Serializar y descargar ────────────────────────────────
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name=f"calificaciones_{course.get('code', 'curso')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"❌ Error al generar Excel: {str(e)}")

def render_settings_tab(conn, course):
    """Renderiza la pestaña de configuración"""
    st.subheader("⚙️ Configuración del Curso")
    
    with st.form("course_settings"):
        new_name = st.text_input("Nombre del curso*", course.get('name', ''))
        new_code = st.text_input("Código*", course.get('code', ''))
        new_desc = st.text_area("Descripción", course.get('description', ''), height=100)
        
        col_status, col_level = st.columns(2)
        new_status = col_status.selectbox(
            "Estado", 
            ["active", "draft", "archived"], 
            index=["active", "draft", "archived"].index(course.get('status', 'active')),
            format_func=lambda x: {
                "active": "🟢 Activo",
                "draft": "📝 Borrador", 
                "archived": "🗄️ Archivado"
            }[x]
        )
        
        # Verificar que 'level' tenga un valor válido, si no, usar 'Básico'
        current_level = course.get('level', 'Básico')
        if current_level not in ["Básico", "Intermedio", "Avanzado"]:
            current_level = "Básico"

        new_level = col_level.selectbox(
            "Nivel", 
            ["Básico", "Intermedio", "Avanzado"],
            index=["Básico", "Intermedio", "Avanzado"].index(current_level)
        )
        
        new_image = st.file_uploader(
            "Portada del curso (opcional)", 
            type=['png', 'jpg', 'jpeg'],
            help="Imagen de portada. Tamaño recomendado: 1200x400px"
        )
        
        col_save, col_cancel = st.columns([2, 1])
        
        if col_save.form_submit_button("💾 Guardar Cambios", type="primary", width='stretch'):
            if not new_name or not new_code:
                st.error("⚠️ Nombre y código son obligatorios")
            else:
                image_blob = new_image.getvalue() if new_image else course.get('cover_image')
                try:
                    conn.execute("""
                        UPDATE courses 
                        SET name=?, code=?, description=?, cover_image=?, status=?, level=?, updated_at=? 
                        WHERE id=?
                    """, (
                        new_name, new_code, new_desc, image_blob, new_status, new_level, 
                        datetime.now(), course['id']
                    ))
                    conn.commit()
                    
                    # Actualizar sesión
                    st.session_state.active_course.update({
                        'name': new_name,
                        'code': new_code,
                        'description': new_desc,
                        'cover_image': image_blob,
                        'status': new_status,
                        'level': new_level
                    })
                    
                    st.success("✅ Cambios guardados exitosamente")
                    time.sleep(1)
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error al guardar: {str(e)}")
        
        if col_cancel.form_submit_button("❌ Cancelar", type="secondary", width='stretch'):
            st.rerun()
    
    st.divider()
    st.markdown("#### ⚠️ Zona de Peligro")
    
    with st.expander("🔥 Eliminar Curso", expanded=False):
        st.warning("""
        **⚠️ ADVERTENCIA CRÍTICA**
        
        Esta acción es **IRREVERSIBLE**. Al eliminar el curso:
        
        - Se eliminarán **TODOS** los módulos, materiales, tareas y exámenes
        - Se perderán **TODAS** las calificaciones y entregas
        - Los estudiantes perderán acceso inmediatamente
        - No podrás recuperar la información
        
        **Esta acción no se puede deshacer.**
        """)
        
        confirm_text = st.text_input(
            "Escribe 'ELIMINAR CURSO' para confirmar:",
            placeholder="ELIMINAR CURSO",
            help="Debes escribir exactamente ELIMINAR CURSO en mayúsculas"
        )
        
        col_confirm, _ = st.columns([1, 2])
        
        if col_confirm.button("🗑️ Eliminar Permanentemente", type="secondary", disabled=True):
            if confirm_text == "ELIMINAR CURSO":
                try:
                    conn.execute("DELETE FROM courses WHERE id = ?", (course['id'],))
                    conn.commit()
                    st.session_state.view_mode = 'dashboard'
                    st.success("✅ Curso eliminado")
                    time.sleep(1.5)
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error al eliminar: {str(e)}")
            else:
                st.error("❌ Texto de confirmación incorrecto")

def view_teacher(conn, model):
    """Vista principal del docente"""
    try:
        # Aquí se ejecuta el chequeo del esquema y se agrega la columna level si falta
        if not check_teacher_schema(conn):
            st.error("❌ Error crítico en la base de datos")
            return
        
        init_teacher_state()
        u = st.session_state.user
        
        # Verificar que el usuario sea docente
        if u.get('role') != 'teacher':
            st.error("❌ Acceso denegado. Solo los docentes pueden acceder a esta vista.")
            st.session_state.view_mode = 'dashboard'
            return
        
        # Manejar página de chat si está activa
        if st.session_state.get('current_page') == 'chat':
            render_chat_interface_teacher(conn, u['username'], u['role'])
            return
        
        # Determinar vista según modo
        if st.session_state.view_mode == 'dashboard':
            render_teacher_dashboard(conn, u)
        elif st.session_state.view_mode == 'course':
            render_teacher_course_view(conn, model, u)
        else:
            st.error("Modo de vista no reconocido")
            st.session_state.view_mode = 'dashboard'
            st.rerun()
            
    except Exception as e:
        st.error("❌ Error crítico en la vista del docente")
        with st.expander("Detalles del error"):
            st.code(traceback.format_exc())
        
        if st.button("🔄 Reiniciar Vista", type="primary"):
            st.session_state.view_mode = 'dashboard'
            st.rerun()



# ==============================================================================
# SISTEMA DE CHAT PRIVADO PARA PROFESORES
# ==============================================================================

def render_chat_interface_teacher(conn, user_id: str, user_role: str):
    """
    Renderiza la interfaz principal del chat privado para profesores.
    Usa las mismas funciones que los estudiantes.
    """
    from utils_chat import chat_manager, format_timestamp
    
    st.title("💬 Chat Privado")
    st.markdown("Comunícate directamente con tus estudiantes")
    
    # Inicializar session state para chat
    if 'selected_contact' not in st.session_state:
        st.session_state.selected_contact = None
    if 'selected_conversation_id' not in st.session_state:
        st.session_state.selected_conversation_id = None
    if 'chat_course_filter' not in st.session_state:
        st.session_state.chat_course_filter = None
    
    # Botón para volver al dashboard
    if st.button("← Volver al Dashboard"):
        st.session_state.current_page = 'dashboard'
        st.session_state.view_mode = 'dashboard'
        st.rerun()
    
    st.markdown("---")
    
    # Layout de dos columnas
    col_contacts, col_chat = st.columns([3, 7])
    
    with col_contacts:
        render_contact_list_teacher(conn, user_id, user_role)
    
    with col_chat:
        if st.session_state.selected_conversation_id:
            render_conversation_window_teacher(conn, st.session_state.selected_conversation_id, user_id)
        else:
            st.info("👈 Selecciona un estudiante para iniciar una conversación")
            st.markdown("""
            ### 💡 Cómo usar el chat
            
            1. **Selecciona un estudiante** de la lista de la izquierda
            2. **Escribe tu mensaje** en el campo de texto
            3. **Adjunta archivos** si lo necesitas (PDF, imágenes, documentos)
            4. **Envía** y recibe respuestas en tiempo real
            
            #### 📋 Características:
            - ✅ Mensajes privados con tus estudiantes
            - 📎 Adjuntar archivos (hasta 10 MB)
            - 🔔 Notificaciones de nuevos mensajes
            - 📊 Contador de mensajes no leídos
            - 🔍 Búsqueda de estudiantes
            """)


def render_contact_list_teacher(conn, user_id: str, user_role: str):
    """
    Renderiza la lista de contactos para profesores.
    """
    from utils_chat import chat_manager, format_timestamp
    
    st.markdown("### 👥 Estudiantes")
    
    # Obtener cursos del profesor
    courses = conn.execute("""
        SELECT id, name FROM courses
        WHERE teacher_id = ?
        ORDER BY name
    """, (user_id,)).fetchall()
    
    # Filtro por curso
    course_options = ["Todos los cursos"] + [f"{c['name']}" for c in courses]
    selected_course_name = st.selectbox("📚 Filtrar por curso:", course_options, key="chat_course_select_teacher")
    
    # Obtener course_id seleccionado
    course_filter = None
    if selected_course_name != "Todos los cursos":
        for c in courses:
            if c['name'] == selected_course_name:
                course_filter = c['id']
                break
    
    # Campo de búsqueda
    search_term = st.text_input("🔍 Buscar estudiante:", placeholder="Nombre del estudiante...", key="chat_search_teacher")
    
    # Obtener contactos
    if search_term:
        contacts = chat_manager.search_contacts(user_id, search_term, course_filter)
    else:
        contacts = chat_manager.get_user_contacts(user_id, course_filter)
    
    st.markdown(f"**{len(contacts)}** estudiante{'s' if len(contacts) != 1 else ''}")
    
    # Acordeón para ver todos los integrantes del curso
    if course_filter:
        with st.expander("🙈 Ver todos los integrantes del curso"):
            # Obtener todos los estudiantes del curso
            students = conn.execute("""
                SELECT u.username, u.first_name, u.last_name, u.avatar, u.role
                FROM users u
                JOIN enrollments e ON u.username = e.student_id
                WHERE e.course_id = ?
                ORDER BY u.last_name, u.first_name
            """, (course_filter,)).fetchall()
            
            # Mostrar estudiantes
            if students:
                st.markdown(f"#### 👨‍🎓 Estudiantes ({len(students)})")
                for student in students:
                    render_all_members_card_teacher(conn, dict(student), user_id, course_filter)
            else:
                st.info("No hay estudiantes inscritos en este curso")
    
    if not contacts:
        st.info("No hay estudiantes disponibles. Usa el acordeón arriba para iniciar un chat con un estudiante del curso.")
        return
    
    # Mostrar contactos
    for contact in contacts:
        render_contact_card_teacher(conn, contact, user_id)


def render_all_members_card_teacher(conn, member: dict, user_id: str, course_id: int):
    """Renderiza una tarjeta de miembro del curso para iniciar chat (profesores)"""
    from utils_chat import chat_manager
    import base64
    
    full_name = f"{member['first_name']} {member['last_name']}"
    role_badge = "👨‍🎓 Estudiante"
    
    # Avatar
    if member.get('avatar'):
        avatar_b64 = base64.b64encode(member['avatar']).decode('utf-8')
        avatar_html = f'<img src="data:image/png;base64,{avatar_b64}" style="width: 48px; height: 48px; border-radius: 50%; object-fit: cover; border: 2px solid #58a6ff;">'
    else:
        initial = member['first_name'][0].upper() if member['first_name'] else 'U'
        color_hash = sum(ord(c) for c in full_name) % 6
        avatar_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']
        avatar_color = avatar_colors[color_hash]
        avatar_html = f'<div style="width: 48px; height: 48px; border-radius: 50%; background-color: {avatar_color}; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 20px; color: white; border: 2px solid #58a6ff;">{initial}</div>'
    
    # Contenedor
    col1, col2, col3 = st.columns([1, 4, 2])
    
    with col1:
        st.markdown(avatar_html, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"**{full_name}**")
        st.caption(role_badge)
    
    with col3:
        if st.button("💬 Chatear", key=f"chat_member_teacher_{member['username']}", use_container_width=True):
            # Crear o obtener conversación
            conversation_id = chat_manager.create_or_get_conversation(
                user_id,
                member['username'],
                course_id
            )
            
            if conversation_id:
                # Configurar el estado correctamente
                st.session_state.selected_conversation_id = conversation_id
                st.session_state.selected_contact = member
                st.rerun()
            else:
                st.error("Error al iniciar conversación")
    
    st.markdown("---")


def render_contact_card_teacher(conn, contact: dict, user_id: str):
    """Renderiza una tarjeta de contacto individual para profesores"""
    from utils_chat import chat_manager, format_timestamp
    import base64
    
    # Determinar si hay mensajes no leídos
    has_unread = contact.get('unread_count', 0) > 0
    
    # Obtener inicial y color del avatar
    contact_name = f"{contact['first_name']} {contact['last_name']}"
    initial = contact['first_name'][0].upper() if contact.get('first_name') else 'U'
    
    # Generar color del avatar basado en el nombre (consistente)
    color_hash = sum(ord(c) for c in contact_name) % 6
    avatar_colors = [
        '#FF6B6B',  # Rojo
        '#4ECDC4',  # Turquesa
        '#45B7D1',  # Azul
        '#FFA07A',  # Naranja
        '#98D8C8',  # Verde agua
        '#F7DC6F'   # Amarillo
    ]
    avatar_color = avatar_colors[color_hash]
    
    # Crear contenedor
    with st.container():
        # Avatar y nombre en columnas
        col_avatar, col_info = st.columns([1, 5])
        
        with col_avatar:
            # Verificar si tiene foto de perfil
            if contact.get('avatar'):
                # Mostrar foto de perfil
                avatar_b64 = base64.b64encode(contact['avatar']).decode()
                st.markdown(f"""
                <div style="
                    width: 48px;
                    height: 48px;
                    border-radius: 50%;
                    overflow: hidden;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.3);
                    margin: 0 auto;
                ">
                    <img src="data:image/png;base64,{avatar_b64}" 
                         style="width: 100%; height: 100%; object-fit: cover;">
                </div>
                """, unsafe_allow_html=True)
            else:
                # Mostrar avatar circular con inicial
                st.markdown(f"""
                <div style="
                    width: 48px;
                    height: 48px;
                    border-radius: 50%;
                    background-color: {avatar_color};
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-weight: bold;
                    font-size: 20px;
                    color: white;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.3);
                    margin: 0 auto;
                ">
                    {initial}
                </div>
                """, unsafe_allow_html=True)
        
        with col_info:
            # Nombre y badge de no leídos
            if has_unread:
                st.markdown(f"**🔴 {contact_name}** ({contact['unread_count']})")
            else:
                st.markdown(f"**{contact_name}**")
            
            # Mostrar curso
            course_name = contact.get('course_name', 'Curso desconocido')
            st.caption(f"� {course_name} • �👤 Estudiante")
            
            # Preview del último mensaje
            if contact.get('last_message_preview'):
                st.caption(f"💬 {contact['last_message_preview'][:40]}...")
            
            # Timestamp
            if contact.get('last_message_at'):
                st.caption(f"🕒 {format_timestamp(contact['last_message_at'])}")
        
        # Botones para abrir chat y borrar
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            if st.button(f"💬 Abrir chat", key=f"contact_teacher_{contact['username']}", use_container_width=True, type="primary" if has_unread else "secondary"):
                # Crear o obtener conversación
                if contact.get('conversation_id'):
                    st.session_state.selected_conversation_id = contact['conversation_id']
                else:
                    # Crear nueva conversación - usar el primer curso del profesor
                    common_course = conn.execute("""
                        SELECT id as course_id FROM courses
                        WHERE teacher_id = ?
                        LIMIT 1
                    """, (user_id,)).fetchone()
                    
                    if common_course:
                        conv_id = chat_manager.create_or_get_conversation(
                            user_id, 
                            contact['username'], 
                            common_course['course_id']
                        )
                        st.session_state.selected_conversation_id = conv_id
                
                st.session_state.selected_contact = contact
                st.rerun()
        
        with col_btn2:
            if st.button(f"🗑️ Borrar", key=f"delete_teacher_{contact['username']}", use_container_width=True, type="secondary"):
                if contact.get('conversation_id'):
                    # Confirmar eliminación
                    if st.session_state.get(f'confirm_delete_teacher_{contact["username"]}'):
                        # Eliminar conversación
                        if chat_manager.delete_conversation(contact['conversation_id'], user_id):
                            # Limpiar estado de confirmación
                            if f'confirm_delete_teacher_{contact["username"]}' in st.session_state:
                                del st.session_state[f'confirm_delete_teacher_{contact["username"]}']
                            st.success("Chat eliminado")
                            st.rerun()
                        else:
                            st.error("Error al eliminar chat")
                    else:
                        # Solicitar confirmación
                        st.session_state[f'confirm_delete_teacher_{contact["username"]}'] = True
                        st.warning("⚠️ Haz clic nuevamente para confirmar")
                        st.rerun()
        
        st.markdown("---")


def render_conversation_window_teacher(conn, conversation_id: int, user_id: str):
    """
    Renderiza la ventana de conversación para profesores.
    Usa la misma lógica que los estudiantes.
    """
    from utils_chat import chat_manager, format_timestamp
    
    # Obtener información del contacto
    contact = st.session_state.get('selected_contact')
    if not contact:
        st.error("Error: No se pudo cargar la conversación")
        return
    
    # Header con nombre del estudiante
    col1, col2 = st.columns([4, 1])
    with col1:
        contact_name = f"{contact['first_name']} {contact['last_name']}"
        st.markdown(f"### 💬 {contact_name}")
        st.caption("Estudiante")
    
    with col2:
        if st.button("✖️ Cerrar", key="close_chat_teacher"):
            st.session_state.selected_conversation_id = None
            st.session_state.selected_contact = None
            st.rerun()
    
    st.markdown("---")
    
    # Obtener historial de mensajes
    messages = chat_manager.get_conversation_history(conversation_id, user_id, limit=50)
    
    # Marcar mensajes como leídos
    chat_manager.mark_messages_as_read(conversation_id, user_id)
    
    # Área de mensajes (invertir orden para mostrar más recientes abajo)
    messages_reversed = list(reversed(messages))
    
    # Contenedor con altura fija y scroll para mensajes
    st.markdown("### 💬 Mensajes")
    
    # Usar st.container con height para crear scroll automático
    messages_container = st.container(height=400)
    
    with messages_container:
        if not messages_reversed:
            st.info("No hay mensajes aún. ¡Inicia la conversación!")
        else:
            for msg in messages_reversed:
                is_own = msg['sender_id'] == user_id
                render_message_bubble_teacher(msg, is_own)
            
            # Agregar un elemento invisible al final para hacer scroll
            st.markdown('<div id="scroll-anchor"></div>', unsafe_allow_html=True)
    
    # Usar components.html para ejecutar JavaScript confiable
    import streamlit.components.v1 as components
    components.html(
        """
        <script>
        window.parent.document.querySelectorAll('[data-testid="stVerticalBlock"]').forEach(function(el) {
            if (el.scrollHeight > el.clientHeight) {
                el.scrollTop = el.scrollHeight;
            }
        });
        </script>
        """,
        height=0,
    )
    
    st.markdown("---")
    
    # Input de mensaje (FUERA del contenedor con scroll - siempre visible)
    st.markdown("### ✍️ Escribir mensaje")
    
    # Inicializar contador de mensajes para limpiar el campo
    if 'message_counter_teacher' not in st.session_state:
        st.session_state.message_counter_teacher = 0
    
    col_input, col_send = st.columns([4, 1])
    
    with col_input:
        message_text = st.text_area(
            "Mensaje:",
            height=100,
            placeholder="Escribe tu mensaje aquí...",
            key=f"chat_message_input_teacher_{st.session_state.message_counter_teacher}",
            label_visibility="collapsed"
        )
    
    # Adjuntar archivo
    uploaded_file = st.file_uploader(
        "📎 Adjuntar archivo (PDF, imágenes, documentos - máx 10MB)",
        type=['pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'txt'],
        key="chat_file_upload_teacher"
    )
    
    with col_send:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        if st.button("📤 Enviar", type="primary", use_container_width=True, key="send_message_btn_teacher"):
            # Validar que haya mensaje O archivo
            if (not message_text or not message_text.strip()) and not uploaded_file:
                st.error("⚠️ Debes escribir un mensaje o adjuntar un archivo")
            else:
                with st.spinner("Enviando mensaje..."):
                    # Si no hay texto, usar un mensaje por defecto para el archivo
                    text_to_send = message_text if message_text and message_text.strip() else "📎 Archivo adjunto"
                    
                    # Enviar mensaje
                    message_id = chat_manager.send_message(
                        conversation_id,
                        user_id,
                        text_to_send
                    )
                    
                    if message_id:
                        # Si hay archivo adjunto, enviarlo
                        if uploaded_file:
                            file_content = uploaded_file.read()
                            file_size = len(file_content)
                            
                            # Validar tamaño
                            if file_size > 10485760:  # 10 MB
                                st.error("⚠️ El archivo excede el tamaño máximo de 10 MB")
                            else:
                                success = chat_manager.send_attachment(
                                    message_id,
                                    uploaded_file.name,
                                    file_content,
                                    uploaded_file.type,
                                    file_size
                                )
                                
                                if not success:
                                    st.error("⚠️ Error al adjuntar el archivo")
                        
                        # Incrementar contador para limpiar el campo
                        st.session_state.message_counter_teacher += 1
                        # Recargar inmediatamente sin mostrar mensaje de éxito
                        st.rerun()
                    else:
                        st.error("⚠️ Error al enviar el mensaje")


def render_message_bubble_teacher(message: dict, is_own_message: bool):
    """
    Renderiza una burbuja de mensaje individual para profesores.
    """
    from utils_chat import format_timestamp, chat_manager
    import html as html_module
    import base64
    
    # Escapar el texto del mensaje para prevenir inyección HTML
    message_text = html_module.escape(message['message_text']) if message['message_text'] else ""
    timestamp = format_timestamp(message['sent_at'])
    
    # Obtener inicial del nombre del remitente
    sender_name = message.get('sender_name', 'Usuario')
    initial = sender_name[0].upper() if sender_name else 'U'
    
    # Generar color del avatar basado en el nombre (consistente)
    color_hash = sum(ord(c) for c in sender_name) % 6
    avatar_colors = [
        '#FF6B6B',  # Rojo
        '#4ECDC4',  # Turquesa
        '#45B7D1',  # Azul
        '#FFA07A',  # Naranja
        '#98D8C8',  # Verde agua
        '#F7DC6F'   # Amarillo
    ]
    avatar_color = avatar_colors[color_hash]
    
    # Colores y alineación según remitente
    if is_own_message:
        # Mensajes propios: azul, alineados a la derecha
        bg_gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"
        align = "flex-end"
        flex_direction = "row-reverse"
        avatar_margin = "0 0 0 10px"
    else:
        # Mensajes recibidos: gris oscuro, alineados a la izquierda
        bg_gradient = "linear-gradient(135deg, #2c3e50 0%, #34495e 100%)"
        align = "flex-start"
        flex_direction = "row"
        avatar_margin = "0 10px 0 0"
    
    # Crear HTML del avatar (foto o inicial)
    try:
        if message.get('sender_avatar') and message['sender_avatar']:
            # Tiene foto de perfil
            avatar_b64 = base64.b64encode(message['sender_avatar']).decode('utf-8')
            avatar_content = f'<img src="data:image/png;base64,{avatar_b64}" style="width: 100%; height: 100%; object-fit: cover;">'
        else:
            # Sin foto, mostrar inicial
            avatar_content = f'<div style="display: flex; align-items: center; justify-content: center; width: 100%; height: 100%; font-weight: bold; font-size: 16px; color: white;">{initial}</div>'
    except Exception as e:
        # Si hay error, usar inicial
        avatar_content = f'<div style="display: flex; align-items: center; justify-content: center; width: 100%; height: 100%; font-weight: bold; font-size: 16px; color: white;">{initial}</div>'
    
    # Generar HTML de adjuntos si existen
    attachments_html = ""
    if message.get('has_attachment') and message.get('attachments'):
        user_id = st.session_state.user['username']
        for attachment in message['attachments']:
            file_data = chat_manager.get_attachment(attachment['id'], user_id)
            if file_data:
                file_type = attachment['file_type']
                
                # Imágenes: mostrar miniatura inline
                if file_type in ['image/png', 'image/jpeg', 'image/gif']:
                    try:
                        img_b64 = base64.b64encode(file_data['file_content']).decode('utf-8')
                        attachments_html += f'<div style="margin-top: 8px;"><img src="data:{file_type};base64,{img_b64}" style="max-width: 200px; max-height: 200px; border-radius: 8px; display: block; cursor: pointer; box-shadow: 0 2px 4px rgba(0,0,0,0.3);" onclick="window.open(this.src, \'_blank\')"></div>'
                    except:
                        pass
                else:
                    # Otros archivos: mostrar icono y nombre con enlace de descarga
                    file_icon = {
                        'application/pdf': '📄',
                        'text/plain': '📃',
                        'application/msword': '📝',
                        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '📝'
                    }.get(file_type, '📎')
                    
                    size_mb = attachment['file_size'] / (1024 * 1024)
                    size_str = f"{size_mb:.2f} MB" if size_mb >= 1 else f"{attachment['file_size'] / 1024:.0f} KB"
                    file_name_short = attachment['file_name'][:30] + "..." if len(attachment['file_name']) > 30 else attachment['file_name']
                    
                    # Crear data URI para descarga
                    file_b64 = base64.b64encode(file_data['file_content']).decode('utf-8')
                    data_uri = f"data:{file_type};base64,{file_b64}"
                    
                    attachments_html += f'<a href="{data_uri}" download="{attachment["file_name"]}" style="text-decoration: none; color: inherit;"><div style="margin-top: 8px; padding: 8px 12px; background: rgba(255,255,255,0.1); border-radius: 8px; display: inline-block; cursor: pointer; transition: background 0.2s;" onmouseover="this.style.background=\'rgba(255,255,255,0.2)\'" onmouseout="this.style.background=\'rgba(255,255,255,0.1)\'"><div style="font-size: 13px;">{file_icon} {file_name_short}</div><div style="font-size: 11px; opacity: 0.7; margin-top: 2px;">{size_str} • Click para descargar</div></div></a>'
    
    # HTML del mensaje con avatar y adjuntos integrados (TODO EN UNA LÍNEA para evitar problemas)
    bubble_html = f'<div style="display: flex; justify-content: {align}; margin-bottom: 15px;"><div style="display: flex; flex-direction: {flex_direction}; align-items: flex-end; max-width: 75%;"><div style="width: 36px; height: 36px; border-radius: 50%; background-color: {avatar_color}; overflow: hidden; flex-shrink: 0; margin: {avatar_margin}; box-shadow: 0 2px 4px rgba(0,0,0,0.2);">{avatar_content}</div><div style="background: {bg_gradient}; color: white; padding: 12px 16px; border-radius: 18px; box-shadow: 0 2px 5px rgba(0,0,0,0.2); word-wrap: break-word;"><div style="font-size: 14px; line-height: 1.4; margin-bottom: 5px;">{message_text}</div>{attachments_html}<div style="font-size: 11px; opacity: 0.7; text-align: right; margin-top: 5px;">{timestamp}</div></div></div></div>'
    
    st.markdown(bubble_html, unsafe_allow_html=True)
